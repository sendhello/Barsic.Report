# -*- coding: utf-8 -*-
#
# This file created with KivyCreatorProject
# <https://github.com/HeaTTheatR/KivyCreatorProgect
#
# Copyright © 2017 Easy
#
# For suggestions and questions:
# <kivydevelopment@gmail.com>
# 
# LICENSE: MIT

import os
import sys
from ast import literal_eval
import logging
from datetime import datetime, timedelta
import pyodbc
from decimal import Decimal
from lxml import etree, objectify
import csv
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Side
from dateutil.relativedelta import relativedelta

from kivy.app import App
from kivy.uix.modalview import ModalView
from kivy.lang import Builder
from kivy.core.window import Window
from kivy.config import ConfigParser
from kivy.clock import Clock
from kivy.utils import get_color_from_hex, get_hex_from_color
from kivy.metrics import dp
from kivy.properties import ObjectProperty, StringProperty
from kivymd.dialog import MDDialog
from kivymd.bottomsheet import MDListBottomSheet, MDGridBottomSheet
from kivymd.textfields import MDTextField

from main import __version__
from libs.translation import Translation
from libs.uix.baseclass.startscreen import StartScreen
from libs.uix.lists import Lists
from libs.utils.showplugins import ShowPlugins

from libs import functions, to_google_sheets

from kivymd.theming import ThemeManager
from kivymd.label import MDLabel
from kivymd.time_picker import MDTimePicker
from kivymd.date_picker import MDDatePicker

from toast import toast
from dialogs import card
import yadisk
import xlwt
import itertools

import urllib
import re
import ntplib
import requests
import webbrowser
import httplib2
import apiclient.discovery
from oauth2client.service_account import ServiceAccountCredentials
import telepot
import socks, socket

logging.basicConfig(filename="barsic_reports.log", level=logging.INFO)

class BarsicReport2(App):
    """
    Функционал программы.
    """

    title = 'Барсик.Отчеты'
    icon = 'icon.png'
    nav_drawer = ObjectProperty()
    theme_cls = ThemeManager()
    theme_cls.primary_palette = 'Purple'
    theme_cls.theme_style = 'Light'
    lang = StringProperty('ru')

    previous_date_from = ObjectProperty()
    previous_date_to = ObjectProperty()

    def __init__(self, **kvargs):
        super(BarsicReport2, self).__init__(**kvargs)
        Window.bind(on_keyboard=self.events_program)
        Window.soft_input_mode = 'below_target'

        self.list_previous_screens = ['base']
        self.window = Window
        self.plugin = ShowPlugins(self)
        self.config = ConfigParser()
        self.manager = None
        self.window_language = None
        self.exit_interval = False
        self.dict_language = literal_eval(
            open(
                os.path.join(self.directory, 'data', 'locales', 'locales.txt')).read()
        )
        self.translation = Translation(
            self.lang, 'Ttest', os.path.join(self.directory, 'data', 'locales')
        )
        self.date_from = datetime.strptime(datetime.now().strftime('%Y-%m-%d'), '%Y-%m-%d')
        self.date_to = self.date_from + timedelta(1)

        self.org1 = ''
        self.org2 = ''

        self.count_sql_error = 0
        self.org_for_finreport = {}
        self.new_service = []
        self.orgs = []
        self.new_agentservice = []
        self.agentorgs = []

    def get_application_config(self):
        return super(BarsicReport2, self).get_application_config(
            '{}/%(appname)s.ini'.format(self.directory))

    def build_config(self, config):
        """Создаёт файл настроек приложения barsicreport2.ini."""
        config.adddefaultsection('General')
        config.setdefault('General', 'language', 'ru')
        config.setdefault('General', 'finreport_xls', 'False')
        config.setdefault('General', 'finreport_google', 'False')
        config.setdefault('General', 'finreport_telegram', 'False')
        config.setdefault('General', 'agentreport_xls', 'False')
        config.setdefault('General', 'split_by_days', 'False')
        config.setdefault('General', 'date_switch', 'True')
        config.setdefault('General', 'use_yadisk', 'False')
        config.setdefault('General', 'check_client_count_total_xls', 'False')
        config.setdefault('General', 'check_cashreport_xls', 'False')
        config.setdefault('General', 'check_itogreport_xls', 'False')
        config.adddefaultsection('MSSQL')
        config.setdefault('MSSQL', 'driver', '{SQL Server}')
        config.setdefault('MSSQL', 'server', '127.0.0.1\\SQLEXPRESS')
        config.setdefault('MSSQL', 'user', 'sa')
        config.setdefault('MSSQL', 'pwd', 'password')
        config.setdefault('MSSQL', 'database1', 'database')
        config.setdefault('MSSQL', 'database2', 'database')
        config.setdefault('MSSQL', 'database_bitrix', 'database')
        config.adddefaultsection('PATH')
        config.setdefault('PATH', 'reportXML', 'data/org_for_report.xml')
        config.setdefault('PATH', 'agentXML', 'data/org_plat_agent.xml')
        config.setdefault('PATH', 'local_folder', 'report')
        config.setdefault('PATH', 'path', 'report')
        config.setdefault('PATH', 'CREDENTIALS_FILE', 'data/1720aecc5640.json')
        config.setdefault('PATH', 'list_google_docs', 'data/list_google_docs.csv')
        config.adddefaultsection('Bitrix')
        config.setdefault('Bitrix', 'bitrix_exchange_url', 'example.site')
        config.setdefault('Bitrix', 'bitrix_exchange_path', '/bitrix/admin/1c_exchange.php')
        config.setdefault('Bitrix', 'bitrix_login', 'login')
        config.setdefault('Bitrix', 'bitrix_password', 'password')
        config.adddefaultsection('Yadisk')
        config.setdefault('Yadisk', 'yadisk_token', 'token')
        config.adddefaultsection('Telegram')
        config.setdefault('Telegram', 'telegram_token', '111111111:aaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa')
        config.setdefault('Telegram', 'telegram_chanel_id', '111111111111')
        config.setdefault('Telegram', 'telegram_proxy_use', 'False')
        config.setdefault('Telegram', 'telegram_proxy_type', '#PROXY_TYPE_SOCKS4 or PROXY_TYPE_SOCKS5 or PROXY_TYPE_HTTP')
        config.setdefault('Telegram', 'telegram_proxy_ip', '127.0.0.1')
        config.setdefault('Telegram', 'telegram_proxy_port', '1080')
        config.setdefault('Telegram', 'telegram_proxy_auth', 'False')
        config.setdefault('Telegram', 'telegram_proxy_username', 'username')
        config.setdefault('Telegram', 'telegram_proxy_password', 'password')
        config.adddefaultsection('GoogleShets')
        config.setdefault('GoogleShets', 'google_all_read', 'False')
        config.setdefault('GoogleShets', 'google_reader_list', '')
        config.setdefault('GoogleShets', 'google_writer_list', '')

    def set_value_from_config(self):
        '''Устанавливает значения переменных из файла настроек barsicreport2.ini.'''
        self.config.read(os.path.join(self.directory, 'barsicreport2.ini'))
        self.lang = self.config.get('General', 'language')
        self.finreport_xls = functions.to_bool(self.config.get('General', 'finreport_xls'))
        self.finreport_google = functions.to_bool(self.config.get('General', 'finreport_google'))
        self.finreport_telegram = functions.to_bool(self.config.get('General', 'finreport_telegram'))
        self.agentreport_xls = functions.to_bool(self.config.get('General', 'agentreport_xls'))
        # self.split_by_days = functions.to_bool(self.config.get('General', 'split_by_days'))
        self.split_by_days = False
        self.date_switch = functions.to_bool(self.config.get('General', 'date_switch'))
        self.use_yadisk = functions.to_bool(self.config.get('General', 'use_yadisk'))
        self.check_client_count_total_xls = functions.to_bool(self.config.get('General', 'check_client_count_total_xls'))
        self.check_cashreport_xls = functions.to_bool(self.config.get('General', 'check_cashreport_xls'))
        self.check_itogreport_xls = functions.to_bool(self.config.get('General', 'check_itogreport_xls'))
        self.driver = self.config.get('MSSQL', 'driver')
        self.server = self.config.get('MSSQL', 'server')
        self.user = self.config.get('MSSQL', 'user')
        self.pwd = self.config.get('MSSQL', 'pwd')
        self.database1 = self.config.get('MSSQL', 'database1')
        self.database2 = self.config.get('MSSQL', 'database2')
        self.database_bitrix = self.config.get('MSSQL', 'database_bitrix')
        self.reportXML = self.config.get('PATH', 'reportXML')
        self.agentXML = self.config.get('PATH', 'agentXML')
        self.local_folder = self.config.get('PATH', 'local_folder')
        self.path = self.config.get('PATH', 'path')
        self.CREDENTIALS_FILE = self.config.get('PATH', 'CREDENTIALS_FILE')
        self.list_google_docs = self.config.get('PATH', 'list_google_docs')
        self.yadisk_token = self.config.get('Yadisk', 'yadisk_token')
        self.bitrix_exchange_url = self.config.get('Bitrix', 'bitrix_exchange_url')
        self.bitrix_exchange_path = self.config.get('Bitrix', 'bitrix_exchange_path')
        self.bitrix_login = self.config.get('Bitrix', 'bitrix_login')
        self.bitrix_password = self.config.get('Bitrix', 'bitrix_password')
        self.telegram_token = self.config.get('Telegram', 'telegram_token')
        self.telegram_chanel_id = self.config.get('Telegram', 'telegram_chanel_id') # '215624388'
        self.telegram_proxy_use = functions.to_bool(self.config.get('Telegram', 'telegram_proxy_use'))
        self.telegram_proxy_type = self.config.get('Telegram', 'telegram_proxy_type')
        self.telegram_proxy_ip = self.config.get('Telegram', 'telegram_proxy_ip')
        self.telegram_proxy_port = self.config.get('Telegram', 'telegram_proxy_port')
        self.telegram_proxy_auth = functions.to_bool(self.config.get('Telegram', 'telegram_proxy_auth'))
        self.telegram_proxy_username = self.config.get('Telegram', 'telegram_proxy_username')
        self.telegram_proxy_password = self.config.get('Telegram', 'telegram_proxy_password')
        self.google_all_read = functions.to_bool(self.config.get('GoogleShets', 'google_all_read'))
        self.google_reader_list = self.config.get('GoogleShets', 'google_reader_list')
        self.google_writer_list = self.config.get('GoogleShets', 'google_writer_list')

    def build(self):
        self.set_value_from_config()
        self.load_all_kv_files(os.path.join(self.directory, 'libs', 'uix', 'kv'))
        self.screen = StartScreen()  # главный экран программы
        self.manager = self.screen.ids.manager
        self.nav_drawer = self.screen.ids.nav_drawer
        return self.screen

    def load_all_kv_files(self, directory_kv_files):
        for kv_file in os.listdir(directory_kv_files):
            kv_file = os.path.join(directory_kv_files, kv_file)
            if os.path.isfile(kv_file):
                with open(kv_file, encoding='utf-8') as kv:
                    Builder.load_string(kv.read())

    def events_program(self, instance, keyboard, keycode, text, modifiers):
        '''Вызывается при нажатии кнопки Меню или Back Key
        на мобильном устройстве.'''

        if keyboard in (1001, 27):
            if self.nav_drawer.state == 'open':
                self.nav_drawer.toggle_nav_drawer()
            self.back_screen(event=keyboard)
        elif keyboard in (282, 319):
            pass

        return True

    def back_screen(self, event=None):
        '''Менеджер экранов. Вызывается при нажатии Back Key
        и шеврона "Назад" в ToolBar.'''

        # Нажата BackKey.
        if event in (1001, 27):
            if self.manager.current == 'base':
                self.dialog_exit()
                return
            try:
                self.manager.current = self.list_previous_screens.pop()
            except:
                self.manager.current = 'base'
            self.screen.ids.action_bar.title = self.title
            self.screen.ids.action_bar.left_action_items = \
                [['menu', lambda x: self.nav_drawer._toggle()]]

    def show_plugins(self, *args):
        '''Выводит на экран список плагинов.'''

        self.plugin.show_plugins()

    def show_about(self, *args):
        self.nav_drawer.toggle_nav_drawer()
        self.screen.ids.about.ids.label.text = \
            self.translation._(
                u'[size=20][b]Барсик.Отчеты[/b][/size]\n\n'
                u'[b]Version:[/b] {version}\n'
                u'[b]License:[/b] Corporate\n\n'
                u'[size=20][b]Developer[/b][/size]\n\n'
                u'[ref=github.com/sendhello]'
                u'[color={link_color}]SendHello[/color][/ref]\n\n'
                u'[b]Source code:[/b] '
                u'[ref=github.com/sendhello/Barsic.Report]'
                u'[color={link_color}]GitHub[/color][/ref]').format(
                version=__version__,
                link_color=get_hex_from_color(self.theme_cls.primary_color)
            )
        self.manager.current = 'about'
        self.screen.ids.action_bar.left_action_items = \
            [['chevron-left', lambda x: self.back_screen(27)]]

    def show_reports(self, *args):
        """
        Переход на экран ОТЧЕТЫ
        :param args:
        :return:
        """
        self.nav_drawer.toggle_nav_drawer()
        self.manager.current = 'report'
        self.screen.ids.action_bar.left_action_items = \
            [['chevron-left', lambda x: self.back_screen(27)]]
        # Загрузка параметров из INI-файла
        self.load_checkbox()
        self.set_date_from(datetime.now().date())

    def show_license(self, *args):
        """
        Переход на экран ЛИЦЕНЗИЯ
        :param args:
        :return:
        """
        self.screen.ids.license.ids.text_license.text = \
            self.translation._('%s') % open(
                os.path.join(self.directory, 'LICENSE'), encoding='utf-8').read()
        self.nav_drawer._toggle()
        self.manager.current = 'license'
        self.screen.ids.action_bar.left_action_items = \
            [['chevron-left', lambda x: self.back_screen(27)]]
        self.screen.ids.action_bar.title = \
            self.translation._('MIT LICENSE')

    def select_locale(self, *args):
        """
        Выводит окно со списком имеющихся языковых локализаций для
        установки языка приложения.
        :param args:
        :return:
        """

        def select_locale(name_locale):
            """
            Устанавливает выбранную локализацию.
            :param name_locale:
            :return:
            """

            for locale in self.dict_language.keys():
                if name_locale == self.dict_language[locale]:
                    self.lang = locale
                    self.config.set('General', 'language', self.lang)
                    self.config.write()

        dict_info_locales = {}
        for locale in self.dict_language.keys():
            dict_info_locales[self.dict_language[locale]] = \
                ['locale', locale == self.lang]

        if not self.window_language:
            self.window_language = card(
                Lists(
                    dict_items=dict_info_locales,
                    events_callback=select_locale, flag='one_select_check'
                ),
                size=(.85, .55)
            )
        self.window_language.open()

    def dialog_exit(self):
        def check_interval_press(interval):
            self.exit_interval += interval
            if self.exit_interval > 5:
                self.exit_interval = False
                Clock.unschedule(check_interval_press)

        if self.exit_interval:
            sys.exit(0)

        Clock.schedule_interval(check_interval_press, 1)
        toast(self.translation._('Press Back to Exit'))

    def show_dialog(self, title, text, func=functions.func_pass, *args, **kwargs):
        content = MDLabel(font_style='Body1',
                          theme_text_color='Secondary',
                          text=text,
                          size_hint_y=None,
                          valign='top')
        content.bind(texture_size=content.setter('size'))
        dialog = MDDialog(title=title,
                               content=content,
                               size_hint=(.8, None),
                               height=dp(200),
                               auto_dismiss=False)

        dialog.add_action_button("Закрыть", action=lambda *x: (dialog.dismiss(), func(*args, **kwargs)))
        dialog.open()

    def show_dialog_variant(self, title, text, func=functions.func_pass, *args, **kwargs):
        content = MDLabel(font_style='Body1',
                          theme_text_color='Secondary',
                          text=text,
                          size_hint_y=None,
                          valign='top')
        content.bind(texture_size=content.setter('size'))
        dialog = MDDialog(title=title,
                               content=content,
                               size_hint=(.8, None),
                               height=dp(200),
                               auto_dismiss=False)

        dialog.add_action_button("ДА", action=lambda *x: (dialog.dismiss(), func(*args, **kwargs)))
        dialog.add_action_button("Нет", action=lambda *x: (dialog.dismiss(), False))
        dialog.open()

    def show_dialog_variant2(self, title, text, func_yes=functions.func_pass, func_no=functions.func_pass):
        content = MDLabel(font_style='Body1',
                          theme_text_color='Secondary',
                          text=text,
                          size_hint_y=None,
                          valign='top')
        content.bind(texture_size=content.setter('size'))
        dialog = MDDialog(title=title,
                               content=content,
                               size_hint=(.8, None),
                               height=dp(200),
                               auto_dismiss=False)

        dialog.add_action_button("ДА", action=lambda *x: (dialog.dismiss(), func_yes))
        dialog.add_action_button("Нет", action=lambda *x: (dialog.dismiss(), func_no))
        dialog.open()

    def on_lang(self, instance, lang):
        self.translation.switch_lang(lang)

    def get_time_picker_data(self, instance, time):
        self.root.ids.time_picker_label.text = str(time)
        self.previous_time = time

    def show_time_picker(self):
        self.time_dialog = MDTimePicker()
        self.time_dialog.bind(time=self.get_time_picker_data)
        if self.root.ids.time_picker_use_previous_time.active:
            try:
                self.time_dialog.set_time(self.previous_time)
            except AttributeError:
                pass
        self.time_dialog.open()

    def set_date_from(self, date_obj):
        self.previous_date_from = date_obj
        self.date_from = datetime.strptime(str(date_obj), '%Y-%m-%d')
        self.root.ids.report.ids.date_from.text = str(date_obj)
        if self.date_to < self.date_from or self.root.ids.report.ids.date_switch.active:
            self.set_date_to(date_obj)
        logging.info(f'{__name__}: {str(datetime.now())[:-7]}:    Установка периода отчета на {self.date_from} - {self.date_to}')

    def show_date_from(self):
        pd = self.previous_date_from
        try:
            MDDatePicker(self.set_date_from,
                         pd.year, pd.month, pd.day).open()
        except AttributeError:
            MDDatePicker(self.set_date_from).open()

    def set_date_to(self, date_obj):
        self.previous_date_to = date_obj
        self.date_to = datetime.strptime(str(date_obj), '%Y-%m-%d') + timedelta(1)
        self.root.ids.report.ids.date_to.text = str(date_obj)
        if self.date_to < self.date_from:
            self.set_date_from(date_obj)
        logging.info(f'{__name__}: {str(datetime.now())[:-7]}:    Установка периода отчета на {self.date_from} - {self.date_to}')

    def show_date_to(self):
        if self.root.ids.report.ids.date_switch.active:
            pass
        else:
            pd = self.previous_date_to
            try:
                MDDatePicker(self.set_date_to,
                             pd.year, pd.month, pd.day).open()
            except AttributeError:
                MDDatePicker(self.set_date_to).open()

    def click_date_switch(self):
        if self.root.ids.report.ids.date_switch.active:
            self.date_switch = True
            self.root.ids.report.ids.label_date.text = 'Дата:'
            self.set_date_to(self.date_from.date() + timedelta(1))
            self.root.ids.report.ids.date_to.theme_text_color = 'Secondary'
            self.root.ids.report.ids.split_by_days.active = False
            self.root.ids.report.ids.split_by_days.disabled = True
            self.root.ids.report.ids.split_by_days_text.theme_text_color = 'Secondary'
            self.change_checkbox('split_by_days', False)
            self.root.ids.report.ids.finreport_google_text.disabled = False
            self.root.ids.report.ids.finreport_google.disabled = False
        else:
            self.date_switch = False
            self.root.ids.report.ids.label_date.text = 'Период:'
            self.root.ids.report.ids.date_to.theme_text_color = 'Primary'
            self.root.ids.report.ids.split_by_days.disabled = False
            self.root.ids.report.ids.split_by_days.active = True
            self.root.ids.report.ids.split_by_days_text.theme_text_color = 'Primary'
            self.change_checkbox('split_by_days', True)

    def count_clients(
            self,
            driver,
            server,
            database,
            uid,
            pwd,
    ):
        """
        Количество человек в зоне
        :return: Количество человек в зоне
        """

        logging.info(f'{__name__}: {str(datetime.now())[:-7]}:    Выполнение функции "count_clients"')

        result = []

        try:
            logging.info(f'{__name__}: {str(datetime.now())[:-7]}:    Попытка соединения с {server}')

            cnxn = pyodbc.connect(
                f'DRIVER={driver};SERVER={server};DATABASE={database};UID={uid};PWD={pwd}')
            cursor = cnxn.cursor()

            cursor.execute("""
                            SELECT
                                [gr].[c1] as [c11],
                                [gr].[StockCategory_Id] as [StockCategory_Id1],
                                [c].[Name],
                                [c].[NN]
                            FROM
                                (
                                    SELECT
                                        [_].[CategoryId] as [StockCategory_Id],
                                        Count(*) as [c1]
                                    FROM
                                        [AccountStock] [_]
                                            INNER JOIN [SuperAccount] [t1] ON [_].[SuperAccountId] = [t1].[SuperAccountId]
                                    WHERE
                                        [_].[StockType] = 41 AND
                                        [t1].[Type] = 0 AND
                                        [_].[Amount] > 0 AND
                                        NOT ([t1].[IsStuff] = 1)
                                    GROUP BY
                                        [_].[CategoryId]
                                ) [gr]
                                    INNER JOIN [Category] [c] ON [gr].[StockCategory_Id] = [c].[CategoryId]
                           """)
            while True:
                row = cursor.fetchone()
                if row:
                    result.append(row)
                else:
                    break
            logging.info(f'{__name__}: {str(datetime.now())[:-7]}:    Результат функции "count_clients": {result}')
            if not result:
                result.append(('Пусто', 488, '', '0003'))

        except pyodbc.OperationalError as e:
            logging.error(f'{__name__}: {str(datetime.now())[:-7]}:    Ошибка {repr(e)}')
            result.append(('Нет данных', 488, 'Ошибка соединения', repr(e)))
            self.show_dialog(f'Ошибка соединения с {server}: {database}', repr(e))
        except pyodbc.ProgrammingError as e:
            logging.error(f'{__name__}: {str(datetime.now())[:-7]}:    Ошибка {repr(e)}')
            result.append(('Нет данных', 488, 'Ошибка соединения', repr(e)))
            self.show_dialog(f'Невозможно открыть {database}', repr(e))
        return result

    def count_clients_print(self):
        in_zone = self.count_clients(
            driver=self.driver,
            server=self.server,
            database=self.database1,
            uid=self.user,
            pwd=self.pwd,
        )
        self.click_select_org()
        try:
            count_clients = int(self.itog_report(
                    server=self.server,
                    database=self.database1,
                    driver=self.driver,
                    user=self.user,
                    pwd=self.pwd,
                    org=self.org1[0],
                    org_name=self.org1[1],
                    date_from=datetime.now(),
                    date_to=datetime.now()+timedelta(1),
                    hide_zeroes='0',
                    hide_internal='1',
                )['Аквазона'][0])
        except KeyError:
            count_clients = 0
        try:
            count_clients_allday = self.reportClientCountTotals(
                server=self.server,
                database=self.database1,
                driver=self.driver,
                user=self.user,
                pwd=self.pwd,
                org=self.org1[0],
                date_from=datetime.now(),
                date_to=datetime.now() + timedelta(1),
            )[0][1]
        except IndexError:
            count_clients_allday = 0

        self.screen.ids.base.ids.count_clients.text = str(count_clients) + ' / ' + str(count_clients_allday)
        self.screen.ids.base.ids.name_zone.text = str(in_zone[len(in_zone) - 1][2])
        self.screen.ids.base.ids.count.text = str(in_zone[len(in_zone) - 1][0])


    # -------------------------------Кнопки вывода списка организаций для выбора----------------------------------------

    # def select_org1(self):
    #     """
    #     Вывод списка организаций
    #     :return:
    #     """
    #     org_list = self.list_organisation(
    #         server=self.server,
    #         database=self.database1,
    #         uid=self.user,
    #         pwd=self.pwd,
    #         driver=self.driver,
    #     )
    #     if org_list:
    #         bs = MDListBottomSheet()
    #         for org in org_list:
    #             bs.add_item(org[2], lambda x: self.click_select_org(org[0], org[2], self.database1), icon='nfc')
    #         bs.open()
    #
    # def select_org2(self):
    #     """
    #     Вывод списка организаций
    #     :return:
    #     """
    #     org_list = self.list_organisation(
    #         server=self.server,
    #         database=self.database2,
    #         uid=self.user,
    #         pwd=self.pwd,
    #         driver=self.driver,
    #     )
    #     if org_list:
    #         bs = MDListBottomSheet()
    #         for org in org_list:
    #             bs.add_item(org[2], lambda x: self.click_select_org(org[0], org[2], self.database2), icon='nfc')
    #         bs.open()
    #
    # def click_select_org(self, id, name, database):
    #     """
    #     Выбор организации из списка и запись ее в переменную
    #     :param id:
    #     :param name:
    #     :param database:
    #     :return:
    #     """
    #     if database == self.database1:
    #         self.org1 = (id, name)
    #         self.screen.ids.report.ids.org1.text = name
    #     elif database == self.database2:
    #         self.org2 = (id, name)
    #         self.screen.ids.report.ids.org2.text = name

    # ---------- Выбор первой организации из списка организаций (Замена кнопкам выбора организаций) --------------------

    def click_select_org(self):
        """
        Выбор первой организации из списка организаций
        """
        org_list1 = self.list_organisation(
                server=self.server,
                database=self.database1,
                uid=self.user,
                pwd=self.pwd,
                driver=self.driver,
            )
        org_list2 = self.list_organisation(
            server=self.server,
            database=self.database2,
            uid=self.user,
            pwd=self.pwd,
            driver=self.driver,
        )
        self.org1 = (org_list1[0][0], org_list1[0][2])
        self.org2 = (org_list2[0][0], org_list2[0][2])
        logging.info(f'{__name__}: {str(datetime.now())[:-7]}:    Выбраны организации {org_list1[0][2]} и {org_list2[0][2]}')

    def list_organisation(self,
                          server,
                          database,
                          uid,
                          pwd,
                          driver,
                          ):
        """
        Функция делает запрос в базу Барс и возвращает список заведенных в базе организаций в виде списка кортежей
        :param server: str - Путь до MS-SQL сервера, например 'SQLEXPRESS\BASE'
        :param database: str - Имя базы данных Барса, например 'SkiBars2'
        :param uid: str - Имя пользователя базы данных, например 'sa'
        :param pwd: str - Пароль пользователя базы данных, например 'pass'
        :return: list = Список организаций, каджая из которых - кортеж с параметрами организации
        """
        logging.info(f'{__name__}: {str(datetime.now())[:-7]}:    Поиск организаций...')
        result = []
        try:
            logging.info(f'{__name__}: {str(datetime.now())[:-7]}:    Попытка соединения с {server}')
            cnxn = pyodbc.connect(f'DRIVER={driver};SERVER={server};DATABASE={database};UID={uid};PWD={pwd}')
            cursor = cnxn.cursor()

            id_type = 1
            cursor.execute(
                f"""
                SELECT
                    SuperAccountId, Type, Descr, CanRegister, CanPass, IsStuff, IsBlocked, BlockReason, DenyReturn, 
                    ClientCategoryId, DiscountCard, PersonalInfoId, Address, Inn, ExternalId, RegisterTime,LastTransactionTime, 
                    LegalEntityRelationTypeId, SellServicePointId, DepositServicePointId, AllowIgnoreStoredPledge, Email, 
                    Latitude, Longitude, Phone, WebSite, TNG_ProfileId
                FROM
                    SuperAccount
                WHERE
                    Type={id_type}
                """)
            while True:
                row = cursor.fetchone()
                if row:
                    result.append(row)
                else:
                    break
        except pyodbc.OperationalError as e:
            logging.error(f'{__name__}: {str(datetime.now())[:-7]}:    Ошибка {repr(e)}')
            self.show_dialog(f'Ошибка соединения с {server}: {database}', repr(e))
        except pyodbc.ProgrammingError as e:
            logging.error(f'{__name__}: {str(datetime.now())[:-7]}:    Ошибка {repr(e)}')
            self.show_dialog(f'Невозможно открыть {database}', repr(e))
        return result

    @functions.to_googleshet
    @functions.add_date
    @functions.add_sum
    @functions.convert_to_dict
    def itog_report(
            self,
            server,
            database,
            driver,
            user,
            pwd,
            org,
            org_name,
            date_from,
            date_to,
            hide_zeroes='0',
            hide_internal='1',
    ):
        """
        Делает запрос в базу Барс и возвращает итоговый отчет за запрашиваемый период
        :param server: str - Путь до MS-SQL сервера, например 'SQLEXPRESS\BASE'
        :param database: str - Имя базы данных Барса, например 'SkiBars2'
        :param uid: str - Имя пользователя базы данных, например 'sa'
        :param pwd: str - Пароль пользователя базы данных, например 'pass'
        :param sa: str - Id организации в Барсе
        :param date_from: str - Начало отчетного периода в формате: 'YYYYMMDD 00:00:00'
        :param date_to:  str - Конец отчетного периода в формате: 'YYYYMMDD 00:00:00'
        :param hide_zeroes: 0 or 1 - Скрывать нулевые позиции?
        :param hide_internal: 0 or 1 - Скрывать внутренние точки обслуживания?
        :return: Итоговый отчет
        """
        cnxn = pyodbc.connect(f'DRIVER={driver};SERVER={server};DATABASE={database};UID={user};PWD={pwd}')
        date_from = date_from.strftime('%Y%m%d 00:00:00')
        date_to = date_to.strftime('%Y%m%d 00:00:00')
        cursor = cnxn.cursor()
        cursor.execute(
            f"exec sp_reportOrganizationTotals_v2 @sa={org},@from='{date_from}',@to='{date_to}',@hideZeroes={hide_zeroes},"
            f"@hideInternal={hide_internal}")
        report = []
        while True:
            row = cursor.fetchone()
            if row:
                report.append(row)
            else:
                break
        report.append((0, 0, 0, 0, org_name, 0, 'Организация', 'Организация'))
        if len(report) > 1:
            logging.info(f'{__name__}: {str(datetime.now())[:-7]}:    Итоговый отчет сформирован ID организации = {org}, '
                         f'Период: {date_from[:8]}-{date_to[:8]}, Скрывать нули = {hide_zeroes}, .'
                         f'Скрывать внутренние точки обслуживания: {hide_internal})')
        return report

    def reportClientCountTotals(
            self,
            server,
            database,
            driver,
            user,
            pwd,
            org,
            date_from,
            date_to,
    ):
        cnxn = pyodbc.connect(f'DRIVER={driver};SERVER={server};DATABASE={database};UID={user};PWD={pwd}')
        date_from = date_from.strftime('%Y%m%d 00:00:00')
        date_to = date_to.strftime('%Y%m%d 00:00:00')
        cursor = cnxn.cursor()
        cursor.execute(
            f"exec sp_reportClientCountTotals @sa={org},@from='{date_from}',@to='{date_to}',@categoryId=0")
        report = []
        while True:
            row = cursor.fetchone()
            if row:
                report.append(row)
            else:
                break
        if len(report) > 1:
            logging.info(f'{__name__}: {str(datetime.now())[:-7]}:    Отчет по количеству посетителей сформирован '
                         f'ID организации = {org}, Период: {date_from[:8]}-{date_to[:8]}')
        return report

    def client_count_totals_period(
            self,
            server,
            database,
            driver,
            user,
            pwd,
            org,
            org_name,
            date_from,
            date_to,
    ):
        """
        Если выбран 1 день возвращает словарь количества людей за текущий месяц,
        если выбран период возвращает словарь количества людей за период
        """
        count = []

        if date_from + timedelta(1) == date_to:
            first_day = datetime.strptime((date_from.strftime('%Y%m') + '01'), '%Y%m%d')
            count.append((org_name, 1))
        else:
            first_day = date_from
            count.append((org_name, 0))
        total = 0
        while first_day < date_to:
            client_count = self.reportClientCountTotals(
                server=server,
                database=database,
                driver=driver,
                user=user,
                pwd=pwd,
                org=org,
                date_from=first_day,
                date_to=first_day + timedelta(1),
            )
            try:
                count.append((client_count[0][0], client_count[0][1]))
                total += client_count[0][1]
            except IndexError:
                count.append((first_day, 0))
            first_day += timedelta(1)
        count.append(('Итого', total))
        return count

    def cash_report_request(
            self,
            server,
            database,
            driver,
            user,
            pwd,
            date_from,
            date_to,
    ):
        """
        Делает запрос в базу Барс и возвращает суммовой отчет за запрашиваемый период
        :param server: str - Путь до MS-SQL сервера, например 'SQLEXPRESS\BASE'
        :param database: str - Имя базы данных Барса, например 'SkiBars2'
        :param uid: str - Имя пользователя базы данных, например 'sa'
        :param pwd: str - Пароль пользователя базы данных, например 'pass'
        :param sa: str - Id организации в Барсе
        :param date_from: str - Начало отчетного периода в формате: 'YYYYMMDD 00:00:00'
        :param date_to:  str - Конец отчетного периода в формате: 'YYYYMMDD 00:00:00'
        :return: Суммовой отчет
        """
        cnxn = pyodbc.connect(f'DRIVER={driver};SERVER={server};DATABASE={database};UID={user};PWD={pwd}')
        date_from = date_from.strftime('%Y%m%d 00:00:00')
        date_to = date_to.strftime('%Y%m%d 00:00:00')
        cursor = cnxn.cursor()
        cursor.execute(
            f"exec sp_reportCashDeskMoney @from='{date_from}', @to='{date_to}'")
        report = []
        while True:
            row = cursor.fetchone()
            if row:
                report.append(row)
            else:
                break
        if len(report) > 1:
            logging.info(f'{__name__}: {str(datetime.now())[:-7]}:    Суммовой отчет сформирован, '
                         f'Период: {date_from[:8]}-{date_to[:8]}')
        return report

    def service_point_request(
            self,
            server,
            database,
            driver,
            user,
            pwd,
    ):
        """
            Делает запрос в базу Барс и возвращает список рабочих мест
            :param server: str - Путь до MS-SQL сервера, например 'SQLEXPRESS\BASE'
            :param database: str - Имя базы данных Барса, например 'SkiBars2'
            :param uid: str - Имя пользователя базы данных, например 'sa'
            :param pwd: str - Пароль пользователя базы данных, например 'pass'
            :param sa: str - Id организации в Барсе
            :return: Суммовой отчет
        """
        cnxn = pyodbc.connect(f'DRIVER={driver};SERVER={server};DATABASE={database};UID={user};PWD={pwd}')
        cursor = cnxn.cursor()
        cursor.execute(
            f"""
            SELECT
                ServicePointId, Name, SuperAccountId, Type, Code, IsInternal
            FROM 
                ServicePoint
            """
        )
        report = []
        while True:
            row = cursor.fetchone()
            if row:
                report.append(row)
            else:
                break
        if len(report) > 1:
            logging.info(f'{__name__}: {str(datetime.now())[:-7]}:    Список рабочих мест сформирован.')
        return report

    def cashdesk_report(
            self,
            server,
            database,
            driver,
            user,
            pwd,
            date_from,
            date_to,
    ):
        """
        Преобразует запросы из базы в суммовой отчет
        :return: dict
        """
        cash_report = self.cash_report_request(
            server=server,
            database=database,
            driver=driver,
            user=user,
            pwd=pwd,
            date_from=date_from,
            date_to=date_to,
        )
        service_point = self.service_point_request(
            server=server,
            database=database,
            driver=driver,
            user=user,
            pwd=pwd,
        )
        service_point_dict = {}
        for point in service_point:
            service_point_dict[point[0]] = (point[1], point[2], point[3], point[4], point[5])
        report = {}
        for line in cash_report:
            report[line[8]] = []
        for line in cash_report:
            report[line[8]].append([service_point_dict[line[0]][0], line[1], line[2], line[3], line[4], line[5], line[6], line[7]])
        all_sum = ['Итого по отчету', Decimal(0.0), Decimal(0.0), Decimal(0.0), Decimal(0.0), Decimal(0.0), Decimal(0.0), Decimal(0.0)]
        for typpe in report:
            type_sum = ['Итого', Decimal(0.0), Decimal(0.0), Decimal(0.0), Decimal(0.0), Decimal(0.0), Decimal(0.0), Decimal(0.0)]
            for line in report[typpe]:
                i = 0
                while True:
                    i += 1
                    try:
                        type_sum[i] += line[i]
                        all_sum[i] += line[i]
                    except IndexError:
                        break
            report[typpe].append(type_sum)
        report['Итого'] = [all_sum]
        report['Дата'] = [[date_from, date_to]]
        if database == self.database1:
            report['Организация'] = [[self.org1[1]]]
        elif database == self.database2:
            report['Организация'] = [[self.org2[1]]]
        return report

    def read_bitrix_base(self,
                         server,
                         database,
                         user,
                         pwd,
                         driver,
                         date_from,
                         date_to,
                         ):
        """
        Функция делает запрос в базу и возвращает список продаж за аказанную дату
        :param server: str - Путь до MS-SQL сервера, например 'SQLEXPRESS\BASE'
        :param database: str - Имя базы данных Барса, например 'SkiBars2'
        :param uid: str - Имя пользователя базы данных, например 'sa'
        :param pwd: str - Пароль пользователя базы данных, например 'pass'
        :param date_from: str - Начало отчетного периода в формате: 'YYYYMMDD 00:00:00'
        :param date_to:  str - Конец отчетного периода в формате: 'YYYYMMDD 00:00:00'
        :return: list = Список организаций, каджая из которых - кортеж с параметрами организации
        """
        logging.info(f'{__name__}: {str(datetime.now())[:-7]}:    Чтение online-продаж...')
        date_from = date_from.strftime("%Y%m%d") + " 00:00:00"
        date_to = date_to.strftime("%Y%m%d") + " 00:00:00"

        cnxn = pyodbc.connect(f'DRIVER={driver};SERVER={server};DATABASE={database};UID={user};PWD={pwd}')
        cursor = cnxn.cursor()
        cursor.execute(
            f"""
            SELECT 
                Id, OrderNumber, ProductId, ProductName, OrderDate, PayDate, Sum, Pay, Status, Client
            FROM 
                Transactions
            WHERE
                (PayDate >= '{date_from}')and(PayDate < '{date_to}')
            """)
        orders = []
        while True:
            row = cursor.fetchone()
            if row:
                orders.append(row)
            else:
                break
        summ = 0
        for order in orders:
            summ += float(order[6])
        return len(orders), summ

    def read_reportgroup(self, XML):
        """
        Чтение XML с привязкой групп услуг к услугам
        :param path:
        :return:
        """
        with open(XML, encoding='utf-8') as f:
            xml = f.read()

        root = objectify.fromstring(xml)
        orgs_dict = {}

        for org in root.UrFace:
            orgs_dict[org.get('Name')] = []
            for serv in org.Services.Service:
                if serv.get('Name') != 'Пустая обязательная категория':
                    orgs_dict[org.get('Name')].append(serv.get('Name'))

        return orgs_dict

    def find_new_service(self, service_dict, orgs_dict):
        """
        Поиск новых услуг и организаций из XML
        :param service_dict: Итоговый отчет
        :param orgs_dict: словарь из XML-файла
        :return:
        """
        servise_set = set()

        for key in orgs_dict:
            for s in orgs_dict[key]:
                servise_set.add(s)

        for org in service_dict:
            if org not in servise_set and org not in self.new_service:
                self.new_service.append(org)
                servise_set.add(org)

        for key in orgs_dict:
            if key not in self.orgs:
                self.orgs.append(key)

    def distibution_service(self):
        """
        Извлекает услугу из списка новых услуг и вызывает список групп до тех пор пока есть новые услуги,
        затем передает управление следующей функции
        """
        if self.new_service:
            service = self.new_service.pop()
            self.viev_orgs(service)
        else:
            self.agentservice()

    def viev_orgs(self, service):
        """
        Выводит всплывающий список групп, при клике на одну из которых услуга указанная в заголовке добавляется в нее
        """
        bs = MDListBottomSheet()
        bs.add_item(f'К какой группе отчета относится услуга "{service}"? (1 из {len(self.new_service) + 1})',
                    lambda x: x)
        for i in range(len(self.orgs)):
            if self.orgs[i] != 'ИТОГО' and self.orgs[i] != 'Депозит' and self.orgs[i] != 'Дата':
                bs.add_item(self.orgs[i], lambda x: self.select_org(service, x.text), icon='nfc')
        bs.add_item(f'Добавить новую группу отчета...',
                    lambda x: self.show_dialog_add_org("Новая группа", "Название новой группы", service))
        bs.open()

    def show_dialog_add_org(self, title, text, service):
        """
        Выводит диалоговое окно с возможность ввода имени новой группы и двумя кнопками
        """
        content = MDTextField(hint_text="Persistent helper text222",
                              helper_text="Text is always here111",
                              helper_text_mode="persistent",
                              text=text,
                              )
        dialog = MDDialog(title=title,
                          content=content,
                          size_hint=(.8, None),
                          height=dp(200),
                          auto_dismiss=False)
        dialog.add_action_button("Отмена", action=lambda *x: (dialog.dismiss(), self.readd_org(service)))
        dialog.add_action_button("Добавить",
                                 action=lambda *x: (dialog.dismiss(), self.create_new_org(dialog.content.text, service)))
        dialog.open()

    def create_new_org(self, name, service):
        """
        Добавляет новую организацию в список организаций self.orgs, словарь self.orgs_dict и XML конфигурацию.
        Возвращает изьятую ранее услугу в список новых услуг с помощью функции self.readd_org
        """
        logging.info(f'{__name__}: {str(datetime.now())[:-7]}:    Добавление новой группы - {name}')
        logging.info(f'{__name__}: {str(datetime.now())[:-7]}:    Добавление услуги {service} в группу {name}')
        self.orgs.append(name)
        self.orgs_dict[name] = []
        self.readd_org(service)
        with open(self.reportXML, encoding='utf-8') as f:
            xml = f.read()
        root = objectify.fromstring(xml)
        #Добавляем новые организации
        new_org = objectify.SubElement(root, "UrFace")
        new_org.set('Name', name)
        new_servs = objectify.SubElement(new_org, 'Services')
        new_serv = objectify.SubElement(new_servs, 'Service')
        new_serv.set('Name', 'Пустая обязательная категория')
        # удаляем аннотации.
        objectify.deannotate(root)
        etree.cleanup_namespaces(root)
        obj_xml = etree.tostring(root,
                                 encoding='utf-8',
                                 pretty_print=True,
                                 xml_declaration=True
                                 )
        # сохраняем данные в файл.
        try:
            with open(self.reportXML, "w", encoding='utf_8_sig') as xml_writer:
                xml_writer.write(obj_xml.decode('utf-8'))
        except IOError:
            pass


    def readd_org(self, service):
        """
        Возвращает изьятую ранее услугу в список новых услуг, затем вызывает функцию распределения
        """
        self.new_service.append(service)
        self.distibution_service()


    def select_org(self, service, org):
        """
        Добавляет услугу в список услуг и вызывает функцию распределения для других услуг
        """
        logging.info(f'{__name__}: {str(datetime.now())[:-7]}:    Добавление услуги {service} в группу {org}')
        self.orgs_dict[org].append(service)
        #Запись новой услуги в XML
        with open(self.reportXML, encoding='utf-8') as f:
            xml = f.read()
        root = objectify.fromstring(xml)
        # Перечисляем существующие организации в файле и добавляем новые строчки
        for x in root.UrFace:
            if x.get('Name') == org:
                Service = objectify.SubElement(x.Services, "Service")
                Service.set("Name", service)
        # удаляем аннотации.
        objectify.deannotate(root)
        etree.cleanup_namespaces(root)
        obj_xml = etree.tostring(root, encoding='utf-8', pretty_print=True, xml_declaration=True)
        # сохраняем данные в файл.
        try:
            with open(self.reportXML, "w", encoding='utf_8_sig') as xml_writer:
                xml_writer.write(obj_xml.decode('utf-8'))
        except IOError:
            pass
        self.distibution_service()

    def agentservice(self):
        self.agent_dict = self.read_reportgroup(self.agentXML)
        self.find_new_agentservice(self.itog_report_org1, self.agent_dict)
        self.distibution_agentservice()

    def distibution_agentservice(self):
        if self.new_agentservice:
            service = self.new_agentservice.pop()
            self.viev_agentorgs(service)
        else:
            self.save_reports()

    def find_new_agentservice(self, service_dict, orgs_dict):
        """
        Поиск новых услуг и организаций из XML
        :param service_dict: Итоговый отчет
        :param orgs_dict: словарь из XML-файла
        :return:
        """
        servise_set = set()

        for key in orgs_dict:
            for s in orgs_dict[key]:
                servise_set.add(s)

        for org in service_dict:
            if org not in servise_set and org not in self.new_agentservice:
                self.new_agentservice.append(org)
                servise_set.add(org)

        for key in orgs_dict:
            if key not in self.agentorgs:
                self.agentorgs.append(key)

    def viev_agentorgs(self, service):
        """
        Выводит всплывающий список организаций,
        при клике на одну из которых услуга указанная в заголовке добавляется в нее
        """
        bs = MDListBottomSheet()
        bs.add_item(f'К какой организации относится услуга "{service}"? (1 из {len(self.new_agentservice) + 1})',
                    lambda x: x)
        for i in range(len(self.agentorgs)):
            if self.agentorgs[i] != 'ИТОГО' and self.agentorgs[i] != 'Депозит' and self.agentorgs[i] != 'Дата':
                bs.add_item(self.agentorgs[i], lambda x: self.select_agentorg(service, x.text), icon='nfc')
        bs.add_item(f'Добавить новую организацию...',
                    lambda x: self.show_dialog_add_agentorg('Наименование организации', 'ООО Рога и Копыта', service))
        bs.open()

    def select_agentorg(self, service, org):
        """
        Добавляет услугу в список услуг и вызывает функцию распределения для других услуг
        """
        self.agent_dict[org].append(service)
        #Запись новой услуги в XML
        with open(self.agentXML, encoding='utf-8') as f:
            xml = f.read()
        root = objectify.fromstring(xml)
        # Перечисляем существующие организации в файле и добавляем новые строчки
        for x in root.UrFace:
            if x.get('Name') == org:
                Service = objectify.SubElement(x.Services, "Service")
                Service.set("Name", service)
        # удаляем аннотации.
        objectify.deannotate(root)
        etree.cleanup_namespaces(root)
        obj_xml = etree.tostring(root, encoding='utf-8', pretty_print=True, xml_declaration=True)
        # сохраняем данные в файл.
        try:
            with open(self.agentXML, "w", encoding='utf_8_sig') as xml_writer:
                xml_writer.write(obj_xml.decode('utf-8'))
        except IOError:
            pass
        self.distibution_agentservice()

    def show_dialog_add_agentorg(self, title, text, service):
        """
        Выводит диалоговое окно с возможность ввода имени новой организыции и двумя кнопками
        """
        content = MDTextField(hint_text="Persistent helper text222",
                              helper_text="Text is always here111",
                              helper_text_mode="persistent",
                              text=text,
                              )
        dialog = MDDialog(title=title,
                          content=content,
                          size_hint=(.8, None),
                          height=dp(200),
                          auto_dismiss=False)
        dialog.add_action_button("Отмена", action=lambda *x: (dialog.dismiss(), self.readd_agentorg(service)))
        dialog.add_action_button("Добавить",
                                 action=lambda *x: (dialog.dismiss(), self.create_new_agentorg(dialog.content.text, service)))
        dialog.open()

    def create_new_agentorg(self, name, service):
        """
        Добавляет новую организацию в список организаций self.orgs, словарь self.orgs_dict и XML конфигурацию.
        Возвращает изьятую ранее услугу в список новых услуг с помощью функции self.readd_org
        """
        self.agentorgs.append(name)
        self.agent_dict[name] = []
        self.readd_agentorg(service)
        with open(self.agentXML, encoding='utf-8') as f:
            xml = f.read()
        root = objectify.fromstring(xml)
        #Добавляем новые организации
        new_org = objectify.SubElement(root, "UrFace")
        new_org.set('Name', name)
        new_servs = objectify.SubElement(new_org, 'Services')
        new_serv = objectify.SubElement(new_servs, 'Service')
        new_serv.set('Name', 'Пустая обязательная категория')
        # удаляем аннотации.
        objectify.deannotate(root)
        etree.cleanup_namespaces(root)
        obj_xml = etree.tostring(root,
                                 encoding='utf-8',
                                 pretty_print=True,
                                 xml_declaration=True
                                 )
        # сохраняем данные в файл.
        try:
            with open(self.agentXML, "w", encoding='utf_8_sig') as xml_writer:
                xml_writer.write(obj_xml.decode('utf-8'))
        except IOError:
            pass


    def readd_agentorg(self, service):
        """
        Возвращает изьятую ранее услугу в список новых услуг, затем вызывает функцию распределения
        """
        self.new_agentservice.append(service)
        self.distibution_agentservice()

    def fin_report(self):
        """
        Форминует финансовый отчет в установленном формате
        :return - dict
        """
        logging.info(f'{__name__}: {str(datetime.now())[:-7]}:    Формирование финансового отчета')
        self.finreport_dict = {}
        is_aquazona = None
        for key in self.orgs_dict:
            if key != 'Не учитывать':
                self.finreport_dict[key] = [0, 0.00]
                for serv in self.orgs_dict[key]:
                    try:
                        if key == 'Нулевые':
                            self.finreport_dict[key][0] += self.itog_report_org1[serv][0]
                            self.finreport_dict[key][1] += self.itog_report_org1[serv][1]
                        elif key == 'Дата':
                            self.finreport_dict[key][0] = self.itog_report_org1[serv][0]
                            self.finreport_dict[key][1] = self.itog_report_org1[serv][1]
                        elif serv == 'Депозит':
                            self.finreport_dict[key][1] += self.itog_report_org1[serv][1]
                        elif serv == 'Аквазона':
                            self.finreport_dict['Кол-во проходов'] = [self.itog_report_org1[serv][0], 0]
                            self.finreport_dict[key][1] += self.itog_report_org1[serv][1]
                            is_aquazona = True
                        elif serv == 'Организация':
                            pass
                        else:
                            self.finreport_dict[key][0] += self.itog_report_org1[serv][0]
                            self.finreport_dict[key][1] += self.itog_report_org1[serv][1]
                    except KeyError:
                        pass
                    except TypeError:
                        pass
        if not is_aquazona:
            self.finreport_dict['Кол-во проходов'] = [0, 0.00]
        self.finreport_dict['Online Продажи'] = list(self.report_bitrix)
        # self.finreport_dict['ИТОГО'][1] -= self.finreport_dict['Депозит'][1]

    def fin_report_lastyear(self):
        """
        Форминует финансовый отчет за предыдущий год в установленном формате
        :return - dict
        """
        logging.info(f'{__name__}: {str(datetime.now())[:-7]}:    Формирование финансового отчета')
        self.finreport_dict_lastyear = {}
        is_aquazona = None
        for key in self.orgs_dict:
            if key != 'Не учитывать':
                self.finreport_dict_lastyear[key] = [0, 0.00]
                for serv in self.orgs_dict[key]:
                    try:
                        if key == 'Нулевые':
                            self.finreport_dict_lastyear[key][0] += self.itog_report_org1_lastyear[serv][0]
                            self.finreport_dict_lastyear[key][1] += self.itog_report_org1_lastyear[serv][1]
                        elif key == 'Дата':
                            self.finreport_dict_lastyear[key][0] = self.itog_report_org1_lastyear[serv][0]
                            self.finreport_dict_lastyear[key][1] = self.itog_report_org1_lastyear[serv][1]
                        elif serv == 'Депозит':
                            self.finreport_dict_lastyear[key][1] += self.itog_report_org1_lastyear[serv][1]
                        elif serv == 'Аквазона':
                            self.finreport_dict_lastyear['Кол-во проходов'] = [self.itog_report_org1_lastyear[serv][0], 0]
                            self.finreport_dict_lastyear[key][1] += self.itog_report_org1_lastyear[serv][1]
                            is_aquazona = True
                        elif serv == 'Организация':
                            pass
                        else:
                            self.finreport_dict_lastyear[key][0] += self.itog_report_org1_lastyear[serv][0]
                            self.finreport_dict_lastyear[key][1] += self.itog_report_org1_lastyear[serv][1]
                    except KeyError:
                        pass
                    except TypeError:
                        pass
        if not is_aquazona:
            self.finreport_dict_lastyear['Кол-во проходов'] = [0, 0.00]
        self.finreport_dict_lastyear['Online Продажи'] = list(self.report_bitrix_lastyear)

    def agent_report(self):
        """
        Форминует отчет платежного агента в установленном формате
        :return - dict
        """
        logging.info(f'{__name__}: {str(datetime.now())[:-7]}:    Формирование отчета платежного агента')
        self.agentreport_dict = {}
        self.agentreport_dict['Организация'] = [self.org1[0], self.org1[1]]
        for key in self.agent_dict:
            if key != 'Не учитывать':
                self.agentreport_dict[key] = [0, 0]
                for serv in self.agent_dict[key]:
                    try:
                        if key == 'Дата':
                            self.agentreport_dict[key][0] = self.itog_report_org1[serv][0]
                            self.agentreport_dict[key][1] = self.itog_report_org1[serv][1]
                        elif serv == 'Депозит':
                            self.agentreport_dict[key][1] += self.itog_report_org1[serv][1]
                        elif serv == 'Аквазона':
                            self.agentreport_dict[key][1] += self.itog_report_org1[serv][1]
                        elif serv == 'Организация':
                            pass
                        else:
                            self.agentreport_dict[key][0] += self.itog_report_org1[serv][0]
                            self.agentreport_dict[key][1] += self.itog_report_org1[serv][1]
                    except KeyError:
                        pass
                    except TypeError:
                        pass

    def export_fin_report(self):
        """
        Сохраняет Финансовый отчет в виде Excel-файла в локальную директорию
        """
        # определяем стили
        h1 = Font(name='Times New Roman',
                  size=18,
                  bold=True,
                  italic=False,
                  vertAlign=None,
                  underline='none',
                  strike=False,
                  color='FF000000')
        font = Font(name='Times New Roman',
                    size=9,
                    bold=False,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='FF000000')
        font_bold = Font(name='Times New Roman',
                         size=9,
                         bold=True,
                         italic=False,
                         vertAlign=None,
                         underline='none',
                         strike=False,
                         color='FF000000')
        fill = PatternFill(fill_type='solid',
                           start_color='c1c1c1',
                           end_color='c2c2c2')
        table_color = PatternFill(fill_type='solid',
                                  start_color='e2e2e2',
                                  end_color='e9e9e9')
        align_top = Alignment(horizontal='general',
                              vertical='top',
                              text_rotation=0,
                              wrap_text=False,
                              shrink_to_fit=False,
                              indent=0,
                              )
        align_top_center = Alignment(horizontal='center',
                              vertical='top',
                              text_rotation=0,
                              wrap_text=False,
                              shrink_to_fit=False,
                              indent=0,
                              )
        border = Border(left=Side(border_style='thin',
                                  color='FF000000'),
                        right=Side(border_style='thin',
                                   color='FF000000'),
                        top=Side(border_style='thin',
                                 color='FF000000'),
                        bottom=Side(border_style='thin',
                                    color='FF000000'),
                        diagonal=Side(border_style='thin',
                                      color='FF000000'),
                        diagonal_direction=0,
                        outline=Side(border_style='thin',
                                     color='FF000000'),
                        vertical=Side(border_style='thin',
                                      color='FF000000'),
                        horizontal=Side(border_style='thin',
                                        color='FF000000')
                        )
        align_left = Alignment(horizontal='left',
                               vertical='bottom',
                               text_rotation=0,
                               wrap_text=False,
                               shrink_to_fit=False,
                               indent=0)
        number_format = 'General'
        protection = Protection(locked=True,
                                hidden=False)

        column = ['', 'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S',
                  'T', 'U', 'V', 'W', 'X', 'Y', 'Z']

        self.row = '0'

        def next_row():
            self.row = str(int(self.row) + 1)
            return self.row

        # объект
        wb = Workbook()

        # активный лист
        ws = wb.active

        # название страницы
        # ws = wb.create_sheet('первая страница', 0)
        ws.title = 'Финансовый отчет'
        # шрифты
        ws['A1'].font = h1
        # выравнивание
        ws['A1'].alignment = align_left

        # Ширина стролбцов
        ws.column_dimensions['A'].width = 1 / 7 * 67
        ws.column_dimensions['B'].width = 1 / 7 * 95
        ws.column_dimensions['C'].width = 1 / 7 * 87
        ws.column_dimensions['D'].width = 1 / 7 * 87
        ws.column_dimensions['E'].width = 1 / 7 * 47
        ws.column_dimensions['F'].width = 1 / 7 * 87
        ws.column_dimensions['G'].width = 1 / 7 * 87
        ws.column_dimensions['H'].width = 1 / 7 * 47
        ws.column_dimensions['I'].width = 1 / 7 * 87
        ws.column_dimensions['J'].width = 1 / 7 * 87
        ws.column_dimensions['K'].width = 1 / 7 * 47
        ws.column_dimensions['L'].width = 1 / 7 * 87
        ws.column_dimensions['M'].width = 1 / 7 * 87
        ws.column_dimensions['N'].width = 1 / 7 * 47
        ws.column_dimensions['O'].width = 1 / 7 * 87
        ws.column_dimensions['P'].width = 1 / 7 * 87
        ws.column_dimensions['Q'].width = 1 / 7 * 47
        ws.column_dimensions['R'].width = 1 / 7 * 87
        ws.column_dimensions['S'].width = 1 / 7 * 87
        ws.column_dimensions['T'].width = 1 / 7 * 47
        ws.column_dimensions['U'].width = 1 / 7 * 87
        ws.column_dimensions['V'].width = 1 / 7 * 47
        ws.column_dimensions['W'].width = 1 / 7 * 87
        ws.column_dimensions['X'].width = 1 / 7 * 87
        ws.column_dimensions['Y'].width = 1 / 7 * 87
        ws.column_dimensions['Z'].width = 1 / 7 * 87

        # значение ячейки
        # ws['A1'] = "Hello!"

        ws[column[1] + next_row()] = 'Финансовый отчет'
        ws.merge_cells(start_row=self.row, start_column=1, end_row=self.row, end_column=len(column) - 1)
        # шрифты
        ws[column[1] + self.row].font = h1
        # выравнивание
        ws[column[1] + self.row].alignment = align_left
        # Высота строк
        ws.row_dimensions[1].height = 24

        ws[column[1] + next_row()] = 'За период с:'
        ws[column[1] + self.row].font = font
        ws[column[1] + self.row].alignment = align_top
        ws[column[2] + self.row] = (self.finreport_dict["Дата"][0]).strftime("%d.%m.%Y")
        ws[column[2] + self.row].font = font_bold
        ws[column[2] + self.row].alignment = align_top
        ws[column[3] + self.row] = 'по'
        ws[column[3] + self.row].font = font
        ws[column[3] + self.row].alignment = align_top
        ws[column[4] + self.row] = (self.finreport_dict["Дата"][1] - timedelta(1)).strftime("%d.%m.%Y")
        ws[column[4] + self.row].font = font_bold
        ws[column[4] + self.row].alignment = align_top

        # ТАБЛИЦА
        self.color = False
        def merge_table():
            # ws.merge_cells(start_row=self.row, start_column=1, end_row=self.row, end_column=4)
            for b in range(1, 27):
                ws[column[b] + self.row].font = font
                ws[column[b] + self.row].alignment = align_top
                if b not in (1, 2, 5, 8, 11, 14, 17, 20, 22):
                    ws[column[b] + self.row].number_format = '#,##0.00 ₽'
                ws[column[b] + self.row].border = border
            if self.color:
                b = 1
                while b < len(column):
                    ws[column[b] + self.row].fill = table_color
                    b += 1
                self.color = False
            else:
                self.color = True

        ws[column[1] + next_row()] = 'Дата'
        ws.merge_cells(start_row=self.row, start_column=1, end_row=str(int(self.row) + 1), end_column=1)
        ws[column[2] + self.row] = 'Кол-во проходов'
        ws.merge_cells(start_row=self.row, start_column=2, end_row=str(int(self.row) + 1), end_column=2)
        ws[column[3] + self.row] = 'Общая сумма'
        ws.merge_cells(start_row=self.row, start_column=3, end_row=str(int(self.row) + 1), end_column=3)
        ws[column[4] + self.row] = 'Сумма KPI'
        ws.merge_cells(start_row=self.row, start_column=4, end_row=str(int(self.row) + 1), end_column=4)
        ws[column[5] + self.row] = 'Билеты'
        ws.merge_cells(start_row=self.row, start_column=5, end_row=self.row, end_column=7)
        ws[column[8] + self.row] = 'Билеты КОРП'
        ws.merge_cells(start_row=self.row, start_column=8, end_row=self.row, end_column=10)
        ws[column[11] + self.row] = 'Термозона'
        ws.merge_cells(start_row=self.row, start_column=11, end_row=self.row, end_column=13)
        ws[column[14] + self.row] = 'Термозона КОРП'
        ws.merge_cells(start_row=self.row, start_column=14, end_row=self.row, end_column=16)
        ws[column[17] + self.row] = 'Общепит'
        ws.merge_cells(start_row=self.row, start_column=17, end_row=self.row, end_column=19)
        ws[column[20] + self.row] = 'Прочее'
        ws.merge_cells(start_row=self.row, start_column=20, end_row=self.row, end_column=21)
        ws[column[22] + self.row] = 'Online Продажи'
        ws.merge_cells(start_row=self.row, start_column=22, end_row=self.row, end_column=24)
        ws[column[25] + self.row] = 'Сумма безнал'
        ws.merge_cells(start_row=self.row, start_column=25, end_row=str(int(self.row) + 1), end_column=25)
        ws[column[26] + self.row] = 'Сумма Biglion'
        ws.merge_cells(start_row=self.row, start_column=26, end_row=str(int(self.row) + 1), end_column=26)
        # раскрвшивание фона для заголовков
        b = 1
        while b < len(column):
            ws[column[b] + self.row].fill = fill
            b += 1
        for b in range(1, 27):
            ws[column[b] + self.row].font = font
            ws[column[b] + self.row].alignment = align_top_center
            ws[column[b] + self.row].border = border

        ws[column[5] + next_row()] = 'Кол-во'
        ws[column[6] + self.row] = 'Сумма'
        ws[column[7] + self.row] = 'Средний чек'
        ws[column[8] + self.row] = 'Кол-во'
        ws[column[9] + self.row] = 'Сумма'
        ws[column[10] + self.row] = 'Средний чек'
        ws[column[11] + self.row] = 'Кол-во'
        ws[column[12] + self.row] = 'Сумма'
        ws[column[13] + self.row] = 'Средний чек'
        ws[column[14] + self.row] = 'Кол-во'
        ws[column[15] + self.row] = 'Сумма'
        ws[column[16] + self.row] = 'Средний чек'
        ws[column[17] + self.row] = 'Кол-во'
        ws[column[18] + self.row] = 'Сумма'
        ws[column[19] + self.row] = 'Средний чек'
        ws[column[20] + self.row] = 'Кол-во'
        ws[column[21] + self.row] = 'Сумма'
        ws[column[22] + self.row] = 'Кол-во'
        ws[column[23] + self.row] = 'Сумма'
        ws[column[24] + self.row] = 'Средний чек'
        # раскрвшивание фона для заголовков
        b = 1
        while b < len(column):
            ws[column[b] + self.row].fill = fill
            b += 1
        for b in range(1, 27):
            ws[column[b] + self.row].font = font
            ws[column[b] + self.row].alignment = align_top_center
            ws[column[b] + self.row].border = border

        if self.finreport_dict['Дата'][0] == self.finreport_dict['Дата'][1] - timedelta(1):
            date_ = datetime.strftime(self.finreport_dict["Дата"][0], "%Y-%m-%d")
        else:
            date_ = f'{datetime.strftime(self.finreport_dict["Дата"][0], "%Y-%m-%d")} - ' \
                    f'{datetime.strftime(self.finreport_dict["Дата"][1] - timedelta(1), "%Y-%m-%d")}'

        ws[column[1] + next_row()] = date_
        ws[column[2] + self.row] = self.finreport_dict['Кол-во проходов'][0]
        ws[column[3] + self.row] = self.finreport_dict['ИТОГО'][1]
        ws[column[4] + self.row] = f'=C{self.row}-U{self.row}+W{self.row}+Y{self.row}+Z{self.row}'
        ws[column[5] + self.row] = self.finreport_dict['Билеты аквапарка'][0]
        ws[column[6] + self.row] = self.finreport_dict['Билеты аквапарка'][1]
        ws[column[7] + self.row] = f'=ЕСЛИОШИБКА(F{self.row}/E{self.row},0)'
        ws[column[8] + self.row] = self.finreport_dict['Билеты аквапарка КОРП'][0]
        ws[column[9] + self.row] = self.finreport_dict['Билеты аквапарка КОРП'][0]
        ws[column[10] + self.row] = f'=ЕСЛИОШИБКА(I{self.row}/H{self.row},0)'
        ws[column[11] + self.row] = self.finreport_dict['Термозона'][0]
        ws[column[12] + self.row] = self.finreport_dict['Термозона'][1]
        ws[column[13] + self.row] = f'=ЕСЛИОШИБКА(L{self.row}/K{self.row},0)'
        ws[column[14] + self.row] = self.finreport_dict['Термозона КОРП'][0]
        ws[column[15] + self.row] = self.finreport_dict['Термозона КОРП'][1]
        ws[column[16] + self.row] = f'=ЕСЛИОШИБКА(O{self.row}/N{self.row},0)'
        ws[column[17] + self.row] = self.finreport_dict['Общепит'][0]
        ws[column[18] + self.row] = self.finreport_dict['Общепит'][1]
        ws[column[19] + self.row] = f'=ЕСЛИОШИБКА(R{self.row}/Q{self.row},0)'
        ws[column[20] + self.row] = self.finreport_dict['Прочее'][0]
        ws[column[21] + self.row] = self.finreport_dict['Прочее'][1]
        ws[column[22] + self.row] = self.finreport_dict['Online Продажи'][0]
        ws[column[23] + self.row] = self.finreport_dict['Online Продажи'][1]
        ws[column[24] + self.row] = f'=ЕСЛИОШИБКА(W{self.row}/V{self.row},0)'
        ws[column[25] + self.row] = 0
        ws[column[26] + self.row] = 0
        merge_table()

        # увеличиваем все строки по высоте
        max_row = ws.max_row
        i = 2
        while i <= max_row:
            rd = ws.row_dimensions[i]
            rd.height = 18
            i += 1
        if self.finreport_dict['Дата'][0] == self.finreport_dict["Дата"][1] - timedelta(1):
            date_ = datetime.strftime(self.finreport_dict["Дата"][0], "%Y-%m-%d")
        else:
            date_ = f'{datetime.strftime(self.finreport_dict["Дата"][0], "%Y-%m-%d")} - ' \
                    f'{datetime.strftime(self.finreport_dict["Дата"][1], "%Y-%m-%d")}'
        path = self.local_folder + self.path + date_ + f' Финансовый отчет' + ".xlsx"
        logging.info(f'{__name__}: {str(datetime.now())[:-7]}:    Сохранение финансового отчета в {path}')
        path = self.create_path(path)
        self.save_file(path, wb)
        return path

    def export_agent_report(self, agentreport_dict):
        """
        Сохраняет отчет платежного агента в виде Excel-файла в локальную директорию
        """
        # определяем стили
        h1 = Font(name='Times New Roman',
                  size=18,
                  bold=True,
                  italic=False,
                  vertAlign=None,
                  underline='none',
                  strike=False,
                  color='FF000000')
        font = Font(name='Times New Roman',
                    size=9,
                    bold=False,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='FF000000')
        font_bold = Font(name='Times New Roman',
                    size=9,
                    bold=True,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='FF000000')
        fill = PatternFill(fill_type='solid',
                           start_color='c1c1c1',
                           end_color='c2c2c2')
        table_color = PatternFill(fill_type='solid',
                                  start_color='e2e2e2',
                                  end_color='e9e9e9')
        align_top = Alignment(horizontal='general',
                              vertical='top',
                              text_rotation=0,
                              wrap_text=False,
                              shrink_to_fit=False,
                              indent=0,
                              )
        border = Border(left=Side(border_style='thin',
                                  color='FF000000'),
                        right=Side(border_style='thin',
                                   color='FF000000'),
                        top=Side(border_style='thin',
                                 color='FF000000'),
                        bottom=Side(border_style='thin',
                                    color='FF000000'),
                        diagonal=Side(border_style='thin',
                                      color='FF000000'),
                        diagonal_direction=0,
                        outline=Side(border_style='thin',
                                     color='FF000000'),
                        vertical=Side(border_style='thin',
                                      color='FF000000'),
                        horizontal=Side(border_style='thin',
                                        color='FF000000')
                        )
        align_left = Alignment(horizontal='left',
                               vertical='bottom',
                               text_rotation=0,
                               wrap_text=False,
                               shrink_to_fit=False,
                               indent=0)
        number_format = 'General'
        protection = Protection(locked=True,
                                hidden=False)

        column = ['', 'A', 'B', 'C', 'D', 'E']

        self.row = '0'
        def next_row():
            self.row = str(int(self.row) + 1)
            return self.row

        # объект
        wb = Workbook()

        # активный лист
        ws = wb.active

        # название страницы
        # ws = wb.create_sheet('первая страница', 0)
        ws.title = 'Отчет платежного агента'
        # шрифты
        ws['A1'].font = h1
        # выравнивание
        ws['A1'].alignment = align_left

        # Ширина стролбцов
        ws.column_dimensions['A'].width = 1 / 7 * 124
        ws.column_dimensions['B'].width = 1 / 7 * 80
        ws.column_dimensions['C'].width = 1 / 7 * 24
        ws.column_dimensions['D'].width = 1 / 7 * 175
        ws.column_dimensions['E'].width = 1 / 7 * 200

        # значение ячейки
        # ws['A1'] = "Hello!"

        ws[column[1] + next_row()] = 'Отчет платежного агента по приему денежных средств'
        ws.merge_cells(start_row=self.row, start_column=1, end_row=self.row, end_column=len(column)-1)
        # шрифты
        ws[column[1] + self.row].font = h1
        # выравнивание
        ws[column[1] + self.row].alignment = align_left
        # Высота строк
        ws.row_dimensions[1].height = 24

        ws[column[1] + next_row()] = f'{agentreport_dict["Организация"][1]}'
        ws.merge_cells(start_row=self.row, start_column=1, end_row=self.row, end_column=len(column) - 1)
        ws[column[1] + self.row].font = font
        ws[column[1] + self.row].alignment = align_top

        ws[column[1] + next_row()] = 'За период с:'
        ws[column[1] + self.row].font = font
        ws[column[1] + self.row].alignment = align_top
        ws[column[2] + self.row] = (agentreport_dict["Дата"][0]).strftime("%d.%m.%Y")
        ws[column[2] + self.row].font = font_bold
        ws[column[2] + self.row].alignment = align_top
        ws[column[3] + self.row] = 'по'
        ws[column[3] + self.row].font = font
        ws[column[3] + self.row].alignment = align_top
        ws[column[4] + self.row] = (agentreport_dict["Дата"][1] - timedelta(1)).strftime("%d.%m.%Y")
        ws[column[4] + self.row].font = font_bold
        ws[column[4] + self.row].alignment = align_top

        # ТАБЛИЦА
        self.color = False
        def merge_table():
            ws.merge_cells(start_row=self.row, start_column=1, end_row=self.row, end_column=4)
            ws[column[1] + self.row].font = font
            ws[column[5] + self.row].font = font
            ws[column[1] + self.row].alignment = align_top
            ws[column[5] + self.row].alignment = align_top
            ws[column[5] + self.row].number_format = '#,##0.00 ₽'
            b = 1
            while b < len(column):
                ws[column[b] + self.row].border = border
                b += 1
            if self.color:
                b = 1
                while b < len(column):
                    ws[column[b] + self.row].fill = table_color
                    b += 1
                self.color = False
            else:
                self.color = True

        def merge_table_bold():
            ws.merge_cells(start_row=self.row, start_column=1, end_row=self.row, end_column=4)
            ws[column[1] + self.row].font = font_bold
            ws[column[5] + self.row].font = font_bold
            ws[column[1] + self.row].alignment = align_top
            ws[column[5] + self.row].alignment = align_top
            b = 1
            while b < len(column):
                ws[column[b] + self.row].border = border
                b += 1

        ws[column[1] + next_row()] = 'Наименование поставщика услуг'
        ws[column[5] + self.row] = 'Сумма'
        merge_table_bold()
        # раскрвшивание фона для заголовков
        b = 1
        while b < len(column):
            ws[column[b] + self.row].fill = fill
            b += 1

        itog_sum = 0
        for line in agentreport_dict:
            if line != 'Организация' and line != 'Дата' and line != 'ИТОГО':
                try:
                    itog_sum += agentreport_dict[line][1]
                    ws[column[1] + next_row()] = line
                    ws[column[5] + self.row] = agentreport_dict[line][1]
                    merge_table()
                except AttributeError:
                    pass

        ws[column[1] + next_row()] = 'Итого'
        if itog_sum != agentreport_dict['ИТОГО'][1]:
            logging.error(f'{__name__}: {str(datetime.now())[:-7]}:    Ошибка. Отчет платежного агента: сумма строк '
                                                                 f'({itog_sum}) не равна строке ИТОГО '
                                                                 f'({agentreport_dict["ИТОГО"][1]})')
            self.show_dialog(f'Ошибка. Отчет платежного агента', f'Ошибка. Отчет платежного агента: сумма строк '
                                                                 f'({itog_sum}) не равна строке ИТОГО '
                                                                 f'({agentreport_dict["ИТОГО"][1]})')
        ws[column[5] + self.row] = itog_sum
        ws[column[5] + self.row].number_format = '#,##0.00 ₽'
        merge_table_bold()

        # увеличиваем все строки по высоте
        max_row = ws.max_row
        i = 2
        while i <= max_row:
            rd = ws.row_dimensions[i]
            rd.height = 18
            i += 1
        if agentreport_dict['Дата'][0] == agentreport_dict["Дата"][1] - timedelta(1):
            date_ = datetime.strftime(agentreport_dict["Дата"][0], "%Y-%m-%d")
        else:
            date_ = f'{datetime.strftime(agentreport_dict["Дата"][0], "%Y-%m-%d")} - ' \
                    f'{datetime.strftime(agentreport_dict["Дата"][1], "%Y-%m-%d")}'
        path = self.local_folder + self.path + date_ + f' Отчет платежного агента {agentreport_dict["Организация"][1]}' + ".xlsx"
        logging.info(f'{__name__}: {str(datetime.now())[:-7]}:    Сохранение отчета платежного агента '
                     f'{agentreport_dict["Организация"][1]} в {path}')
        path = self.create_path(path)
        self.save_file(path, wb)
        return path

    def create_path(self, path):
        """
        Проверяет наличие указанного пути. В случае отсутствия каких-либо папок создает их
        """
        logging.info(f'{__name__}: {str(datetime.now())[:-7]}:    Проверка локальных путей сохранения файлов...')
        list_path = path.split('/')
        path = ''
        end_path = ''
        if list_path[-1][-4:] == '.xls' or list_path[-1]:
            end_path = list_path.pop()
        list_path.append(self.date_from.strftime('%Y'))
        list_path.append(self.date_from.strftime('%m') + '-' + self.date_from.strftime('%B'))
        directory = os.getcwd()
        for folder in list_path:
            if folder not in os.listdir():
                os.mkdir(folder)
                logging.warning(f'{__name__}: {str(datetime.now())[:-7]}:    В директории "{os.getcwd()}" создана папка "{folder}"')
                os.chdir(folder)
            else:
                os.chdir(folder)
            path += folder + '/'
        path += end_path
        os.chdir(directory)
        return path

    def save_file(self, path, file):
        """
        Проверяет не занят ли файл другим процессом и если нет, то перезаписывает его, в противном
        случае выводит диалоговое окно с предложением закрыть файл и продолжить
        """
        try:
            file.save(path)
        except PermissionError as e:
            logging.error(f'{__name__}: {str(datetime.now())[:-7]}:    Файл "{path}" занят другим процессом.\n{repr(e)}')
            self.show_dialog(f'Ошибка записи файла',
                             f'Файл "{path}" занят другим процессом.\nДля повтора попытки закройте это сообщение',
                             func=self.save_file, path=path, file=file)

    def sync_to_yadisk(self, path_list, token):
        """
        Копирует локальные файлы в Яндекс Диск
        """
        logging.info(f'{__name__}: {str(datetime.now())[:-7]}:    Копирование отчетов в Яндекс.Диск...')
        if path_list:
            if self.use_yadisk:
                logging.info(f'{__name__}: {str(datetime.now())[:-7]}:    Соединение с YaDisk...')
                self.yadisk = yadisk.YaDisk(token=token)
                if self.yadisk.check_token():
                    path = '' + self.path
                    remote_folder = self.create_path_yadisk(path)
                    for local_path in path_list:
                        remote_path = remote_folder + local_path.split('/')[-1]
                        logging.info(f'{__name__}: {str(datetime.now())[:-7]}:    Отправка файла "{local_path.split("/")[-1]}" в YaDisk...')
                        files_list_yandex = list(self.yadisk.listdir(remote_folder))
                        files_list = []
                        for key in files_list_yandex:
                            if key['file']:
                                files_list.append(remote_folder + key['name'])
                        if remote_path in files_list:
                            logging.warning(
                                f'{__name__}: {str(datetime.now())[:-7]}:    '
                                f'Файл "{local_path.split("/")[-1]}" уже существует в "{remote_folder}" и будет заменен!')
                            self.yadisk.remove(remote_path, permanently=True)
                        self.yadisk.upload(local_path, remote_path)
                        logging.info(
                            f'{__name__}: {str(datetime.now())[:-7]}:    '
                            f'Файл "{local_path.split("/")[-1]}" отправлен в "{remote_folder}" YaDisk...')
                else:
                    logging.error(f'{__name__}: {str(datetime.now())[:-7]}:    Ошибка YaDisk: token не валиден')
                    self.show_dialog('Ошибка соединения с Yandex.Disc',
                                     f'\nОтчеты сохранены в папке {self.local_folder} '
                                     f'и не будут отправлены на Yandex.Disc.')
        else:
            logging.info(f'{__name__}: {str(datetime.now())[:-7]}:    Нет ни одного отчета для отправки в Yandex.Disk')

    def create_path_yadisk(self, path):
        """
        Проверяет наличие указанного пути в Яндекс Диске. В случае отсутствия каких-либо папок создает их
        :param path:
        :return:
        """
        logging.info(f'{__name__}: {str(datetime.now())[:-7]}:    Проверка путей сохранения файлов на Яндекс.Диске...')
        list_path = path.split('/')
        path = ''
        end_path = ''
        if list_path[-1][-4:] == '.xls' or list_path[-1] == '':
            end_path = list_path.pop()
        list_path.append(self.date_from.strftime('%Y'))
        list_path.append(self.date_from.strftime('%m') + '-' + self.date_from.strftime('%B'))
        directory = '/'
        list_path_yandex = []
        for folder in list_path:
            folder = directory + folder
            directory = folder + '/'
            list_path_yandex.append(folder)
        directory = '/'
        for folder in list_path_yandex:
            folders_list = []
            folders_list_yandex = list(self.yadisk.listdir(directory))
            for key in folders_list_yandex:
                if not key['file']:
                    folders_list.append(directory + key['name'])
            if folder not in folders_list:
                self.yadisk.mkdir(folder)
                logging.info(f'{__name__}: {str(datetime.now())[:-7]}:    Создание новой папки в YandexDisk - "{folder}"')
                directory = folder + '/'
            else:
                directory = folder + '/'
        path = list_path_yandex[-1] + '/'
        return path

    def export_to_google_sheet(self):
        """
        Формирование и заполнение google-таблицы
        """
        logging.info(f'{__name__}: {str(datetime.now())[:-7]}:    Сохранение Финансового отчета в Google-таблицах...')

        #self.CREDENTIALS_FILE # имя файла с закрытым ключом

        self.sheet_width = 37
        height = 34

        credentials = ServiceAccountCredentials.from_json_keyfile_name(self.CREDENTIALS_FILE,
                                                                       ['https://www.googleapis.com/auth/spreadsheets',
                                                                        'https://www.googleapis.com/auth/drive'])
        httpAuth = credentials.authorize(httplib2.Http())
        try:
            self.googleservice = apiclient.discovery.build('sheets', 'v4', http=httpAuth)
        except IndexError as e:
            logging.error(f'{__name__}: {str(datetime.now())[:-7]}:    Ошибка {repr(e)}')

        data_report = datetime.strftime(self.finreport_dict['Дата'][0], '%m')
        month = [
            '',
            'Январь',
            'Февраль',
            'Март',
            'Апрель',
            'Май',
            'Июнь',
            'Июль',
            'Август',
            'Сентябрь',
            'Октябрь',
            'Ноябрь',
            'Декабрь',
        ]
        data_report = month[int(data_report)]

        doc_name = f"{datetime.strftime(self.finreport_dict['Дата'][0], '%Y-%m')} " \
            f"({data_report}) - Финансовый отчет по Аквапарку"

        if self.finreport_dict['Дата'][0] + timedelta(1) != self.finreport_dict['Дата'][1]:
            logging.info(f'{__name__}: {str(datetime.now())[:-7]}:    '
                         f'Экспорт отчета в Google Sheet за несколько дней невозможен!')
            self.show_dialog('Ошибка экспорта в Google.Sheet',
                             'Экспорт отчета в Google Sheet за несколько дней невозможен!')
        else:
            with open(self.list_google_docs, 'r', encoding='utf-8') as f:
                links = csv.reader(f, delimiter=';')
                self.google_links = {}
                for line in links:
                    self.google_links[line[0]] = line[1]
            if self.date_from.strftime('%Y-%m') in self.google_links:
                self.google_doc = (self.date_from.strftime('%Y-%m'),
                                   self.google_links[self.date_from.strftime('%Y-%m')])
            else:
                self.google_doc = None
                # Создание документа
                self.spreadsheet = self.googleservice.spreadsheets().create(body={
                    'properties': {'title': doc_name, 'locale': 'ru_RU'},
                    'sheets': [{'properties': {'sheetType': 'GRID',
                                               'sheetId': 0,
                                               'title': 'Сводный',
                                               'gridProperties': {'rowCount': height, 'columnCount': self.sheet_width}}},
                               {'properties': {'sheetType': 'GRID',
                                               'sheetId': 1,
                                               'title': 'Расширенный',
                                               'gridProperties': {'rowCount': height, 'columnCount': 100}}}
                               ]
                }).execute()

                # Доступы к документу
                self.google_reader_list = self.google_reader_list.split(',')
                self.google_writer_list = self.google_writer_list.split(',')
                driveService = apiclient.discovery.build('drive', 'v3', http=httpAuth)
                if self.google_all_read:
                    shareRes = driveService.permissions().create(
                        fileId=self.spreadsheet['spreadsheetId'],
                        body={'type': 'anyone', 'role': 'reader'},  # доступ на чтение кому угодно
                        fields='id'
                    ).execute()
                # Возможные значения writer, commenter, reader
                # доступ на Чтение определенным пользователоям
                for adress in self.google_reader_list:
                    shareRes = driveService.permissions().create(
                        fileId=self.spreadsheet['spreadsheetId'],
                        body={'type': 'user', 'role': 'reader', 'emailAddress': adress},
                        fields='id'
                    ).execute()
                # доступ на Запись определенным пользователоям
                for adress in self.google_writer_list:
                    shareRes = driveService.permissions().create(
                        fileId=self.spreadsheet['spreadsheetId'],
                        body={'type': 'user', 'role': 'writer', 'emailAddress': adress},
                        fields='id'
                    ).execute()

                sheetId = 0
                # Ширина столбцов
                ss = to_google_sheets.Spreadsheet(self.spreadsheet['spreadsheetId'], sheetId,
                                                  self.googleservice,
                                                  self.spreadsheet['sheets'][sheetId]['properties']['title'])
                ss.prepare_setColumnsWidth(0, 1, 100)
                ss.prepare_setColumnsWidth(2, 7, 120)
                ss.prepare_setColumnWidth(8, 65)
                ss.prepare_setColumnWidth(9, 120)
                ss.prepare_setColumnWidth(10, 100)
                ss.prepare_setColumnWidth(11, 100)
                ss.prepare_setColumnWidth(12, 65)
                ss.prepare_setColumnWidth(13, 120)
                ss.prepare_setColumnWidth(14, 100)
                ss.prepare_setColumnWidth(15, 65)
                ss.prepare_setColumnWidth(16, 120)
                ss.prepare_setColumnWidth(17, 100)
                ss.prepare_setColumnWidth(18, 65)
                ss.prepare_setColumnWidth(19, 120)
                ss.prepare_setColumnWidth(20, 100)
                ss.prepare_setColumnWidth(21, 65)
                ss.prepare_setColumnWidth(22, 120)
                ss.prepare_setColumnWidth(23, 100)
                ss.prepare_setColumnWidth(24, 65)
                ss.prepare_setColumnWidth(25, 120)
                ss.prepare_setColumnWidth(26, 100)
                ss.prepare_setColumnWidth(27, 65)
                ss.prepare_setColumnWidth(28, 120)
                ss.prepare_setColumnWidth(29, 100)
                ss.prepare_setColumnWidth(30, 65)
                ss.prepare_setColumnWidth(31, 120)
                ss.prepare_setColumnWidth(32, 65)
                ss.prepare_setColumnWidth(33, 120)
                ss.prepare_setColumnWidth(34, 100)
                ss.prepare_setColumnWidth(35, 120)
                ss.prepare_setColumnWidth(36, 120)

                # Объединение ячеек
                ss.prepare_mergeCells("A1:A2")
                ss.prepare_mergeCells("B1:B2")
                ss.prepare_mergeCells("C1:C2")
                ss.prepare_mergeCells("D1:D2")
                ss.prepare_mergeCells("E1:E2")
                ss.prepare_mergeCells("F1:F2")
                ss.prepare_mergeCells("G1:G2")
                ss.prepare_mergeCells("H1:H2")
                ss.prepare_mergeCells("I1:K1")
                ss.prepare_mergeCells("L1:L2")
                ss.prepare_mergeCells("M1:O1")
                ss.prepare_mergeCells("P1:R1")
                ss.prepare_mergeCells("S1:U1")
                ss.prepare_mergeCells("V1:X1")
                ss.prepare_mergeCells("Y1:AA1")
                ss.prepare_mergeCells("AB1:AD1")
                ss.prepare_mergeCells("AE1:AF1")
                ss.prepare_mergeCells("AG1:AI1")
                ss.prepare_mergeCells("AJ1:AJ2")
                ss.prepare_mergeCells("AK1:AK2")

                # Задание параметров группе ячеек
                # Жирный, по центру
                ss.prepare_setCellsFormat('A1:AK2', {'horizontalAlignment': 'CENTER', 'textFormat': {'bold': True}})
                # ss.prepare_setCellsFormat('E4:E8', {'numberFormat': {'pattern': '[h]:mm:ss', 'type': 'TIME'}},
                #                           fields='userEnteredFormat.numberFormat')

                # Заполнение таблицы
                ss.prepare_setValues("A1:AK2", [
                    [
                        "Дата", "День недели", "Кол-во проходов \nФАКТ", "Кол-во проходов \nПЛАН",
                    f"Кол-во проходов \n{data_report} "
                    f"{datetime.strftime(self.finreport_dict['Дата'][0] - relativedelta(years=1), '%Y')}",
                    "Общая сумма \nФАКТ", "Общая сумма \nПЛАН",
                    f"Общая сумма \n{data_report} "
                    f"{datetime.strftime(self.finreport_dict['Дата'][0] - relativedelta(years=1), '%Y')}",
                    "Билеты", "", "", "Депозит", "Термозона", "", "", "Общепит", "", "", "Общепит ПЛАН",  "", "",
                    f"Общепит {data_report} "
                    f"{datetime.strftime(self.finreport_dict['Дата'][0] - relativedelta(years=1), '%Y')}", "", "",
                    "Билеты КОРП", "", "", "Термозона КОРП", "", "", "Прочее", "", "Online Продажи", "", "",
                    "Сумма безнал", "Сумма Biglion"
                    ],
                    [
                        "", "", "", "", "", "", "", "", "Кол-во", "Сумма", "Средний чек", "",
                        "Кол-во", "Сумма", "Средний чек", "Кол-во", "Сумма", "Средний чек",
                        "Кол-во", "Сумма", "Средний чек", "Кол-во", "Сумма", "Средний чек",
                        "Кол-во", "Сумма", "Средний чек", "Кол-во", "Сумма", "Средний чек",
                        "Кол-во", "Сумма", "Кол-во", "Сумма", "Средний чек", "", ""
                    ]
                ],
                                     "ROWS")
                # ss.prepare_setValues("D5:E6", [["This is D5", "This is D6"], ["This is E5", "=5+5"]], "COLUMNS")

                # Цвет фона ячеек
                ss.prepare_setCellsFormat("A1:AK2", {"backgroundColor": functions.htmlColorToJSON("#f7cb4d")},
                                          fields="userEnteredFormat.backgroundColor")

                # Бордер
                for i in range(2):
                    for j in range(self.sheet_width):
                        ss.requests.append({"updateBorders": {
                            "range": {"sheetId": ss.sheetId, "startRowIndex": i, "endRowIndex": i + 1,
                                      "startColumnIndex": j,
                                      "endColumnIndex": j + 1},
                            "top": {"style": "SOLID", "width": 1, "color": {"red": 0, "green": 0, "blue": 0}}}})
                        ss.requests.append({"updateBorders": {
                            "range": {"sheetId": ss.sheetId, "startRowIndex": i, "endRowIndex": i + 1,
                                      "startColumnIndex": j,
                                      "endColumnIndex": j + 1},
                            "right": {"style": "SOLID", "width": 1,
                                      "color": {"red": 0, "green": 0, "blue": 0, "alpha": 1.0}}}})
                        ss.requests.append({"updateBorders": {
                            "range": {"sheetId": ss.sheetId, "startRowIndex": i, "endRowIndex": i + 1,
                                      "startColumnIndex": j,
                                      "endColumnIndex": j + 1},
                            "left": {"style": "SOLID", "width": 1,
                                      "color": {"red": 0, "green": 0, "blue": 0, "alpha": 1.0}}}})

                ss.runPrepared()

                self.google_doc = (self.date_from.strftime('%Y-%m'), self.spreadsheet['spreadsheetId'])
                self.google_links[self.google_doc[0]] = self.google_doc[1]
                links = []
                for docid in self.google_links:
                    links.append([docid, self.google_links[docid]])
                with open(self.list_google_docs, 'w', newline='', encoding='utf-8') as f:
                    file = csv.writer(f, delimiter=';')
                    for link in links:
                        file.writerow(link)
                logging.info(
                    f'{__name__}: {str(datetime.now())[:-7]}:    Создана новая таблица с Id: {self.spreadsheet["spreadsheetId"]}')

            self.spreadsheet = self.googleservice.spreadsheets().get(spreadsheetId=self.google_doc[1], ranges=[],
                                                     includeGridData=True).execute()

            # -------------------------------- ЗАПОЛНЕНИЕ ДАННЫМИ ------------------------------------------------

            # Печать таблицы в консоль
            # s = ''
            # for line_table in spreadsheet['sheets'][0]['data'][0]['rowData']:
            #     for cell in line_table['values']:
            #         try:
            #             s += cell['formattedValue'] + " | "
            #         except KeyError:
            #             pass
            #     s = ''

            # Проверка нет ли текущей даты в таблице
            self.nex_line = 1
            self.reprint = 2

            for line_table in self.spreadsheet['sheets'][0]['data'][0]['rowData']:
                try:
                    if line_table['values'][0]['formattedValue'] == datetime.strftime(self.finreport_dict['Дата'][0],
                                                                                      '%d.%m.%Y'):
                        if self.root.ids.report.ids.split_by_days.active:
                            self.rewrite_google_sheet()
                        else:
                            self.show_dialog_variant(f'Перезаписать эту строку?',
                                                     f'Строка за '
                                                     f'{datetime.strftime(self.finreport_dict["Дата"][0], "%d.%m.%Y")} '
                                                     f'уже существует в таблице!',
                                                     self.rewrite_google_sheet,
                                                     )
                        self.reprint = 0
                        break
                    elif line_table['values'][0]['formattedValue'] == "ИТОГО":
                        break
                    else:
                        self.nex_line += 1
                except KeyError:
                    self.nex_line += 1
            if self.reprint:
                self.write_google_sheet()

            # width_table = len(self.spreadsheet['sheets'][0]['data'][0]['rowData'][0]['values'])

    def rewrite_google_sheet(self):
        """
        Заполнение google-таблицы в случае, если данные уже существуют
        """
        logging.warning(f'{__name__}: {str(datetime.now())[:-7]}:    Перезапись уже существующей строки...')
        self.reprint = 1
        self.write_google_sheet()

    def write_google_sheet(self):
        """
        Заполнение google-таблицы
        """
        sheetId = 0
        ss = to_google_sheets.Spreadsheet(self.spreadsheet['spreadsheetId'], sheetId, self.googleservice,
                                          self.spreadsheet['sheets'][sheetId]['properties']['title'])

        # Заполнение строки с данными
        weekday_rus = [
            "Понедельник",
            "Вторник",
            "Среда",
            "Четверг",
            "Пятница",
            "Суббота",
            "Воскресенье",
        ]

        if self.finreport_dict['ИТОГО'][1] != (self.finreport_dict['Билеты аквапарка'][1] +
            self.finreport_dict['Термозона'][1] + self.finreport_dict['Общепит'][1] +
            self.finreport_dict['Билеты аквапарка КОРП'][1] + self.finreport_dict['Прочее'][1] +
            self.finreport_dict['Термозона КОРП'][1] + self.finreport_dict['Депозит'][1]):
            logging.error(f'{__name__}: {str(datetime.now())[:-7]}:    '
                         f'Несоответствие данных: Сумма услуг не равна итоговой сумме')
            self.show_dialog('Несоответствие данных',
                             f"Сумма услуг по группам + депозит ({self.finreport_dict['Билеты аквапарка'][1] + self.finreport_dict['Термозона'][1] + self.finreport_dict['Общепит'][1] + self.finreport_dict['Билеты аквапарка КОРП'][1] + self.finreport_dict['Прочее'][1] + self.finreport_dict['Термозона КОРП'][1] +  + self.finreport_dict['Депозит'][1]}) "
                             f"не равна итоговой сумме ({self.finreport_dict['ИТОГО'][1]}). \n"
                             f"Нулевые позиции имеют стоимость {self.finreport_dict['Нулевые'][1]}. \n"
                             f"Рекомендуется проверить правильно ли разделены услуги по группам.")

        ss.prepare_setValues(
            f"A{self.nex_line}:AK{self.nex_line}",
            [
                [
                    datetime.strftime(self.finreport_dict['Дата'][0], '%d.%m.%Y'),
                    weekday_rus[self.finreport_dict['Дата'][0].weekday()],
                    f"{self.finreport_dict['Кол-во проходов'][0]}",
                    "",
                    f"{self.finreport_dict_lastyear['Кол-во проходов'][0]}",
                    f"={str(self.finreport_dict['ИТОГО'][1]).replace('.', ',')}+AG{self.nex_line}+"
                        f"AI{self.nex_line}+AJ{self.nex_line}",
                    "",
                    f"={str(self.finreport_dict_lastyear['ИТОГО'][1]).replace('.', ',')}+"
                        f"{str(self.finreport_dict_lastyear['Online Продажи'][1]).replace('.', ',')}",
                    self.finreport_dict['Билеты аквапарка'][0],
                    self.finreport_dict['Билеты аквапарка'][1],
                    f"=IFERROR(J{self.nex_line}/I{self.nex_line};0)",
                    self.finreport_dict['Депозит'][1],
                    self.finreport_dict['Термозона'][0],
                    self.finreport_dict['Термозона'][1],
                    f"=IFERROR(M{self.nex_line}/L{self.nex_line};0)",
                    self.finreport_dict['Общепит'][0],
                    self.finreport_dict['Общепит'][1],
                    f"=IFERROR(P{self.nex_line}/O{self.nex_line};0)",
                    "",
                    "",
                    f"=IFERROR(S{self.nex_line}/R{self.nex_line};0)",
                    self.finreport_dict_lastyear['Общепит'][0],
                    self.finreport_dict_lastyear['Общепит'][1],
                    f"=IFERROR(V{self.nex_line}/U{self.nex_line};0)",
                    self.finreport_dict['Билеты аквапарка КОРП'][0],
                    self.finreport_dict['Билеты аквапарка КОРП'][1],
                    f"=IFERROR(Y{self.nex_line}/X{self.nex_line};0)",
                    self.finreport_dict['Термозона КОРП'][0],
                    self.finreport_dict['Термозона КОРП'][1],
                    f"=IFERROR(AB{self.nex_line}/AA{self.nex_line};0)",
                    self.finreport_dict['Прочее'][0],
                    self.finreport_dict['Прочее'][1],
                    self.finreport_dict['Online Продажи'][0],
                    self.finreport_dict['Online Продажи'][1],
                    f"=IFERROR(AG{self.nex_line}/AF{self.nex_line};0)",
                    0,
                    0,
                ]
            ],
            "ROWS"
        )

        # Задание форматы вывода строки
        ss.prepare_setCellsFormats(
            f"A{self.nex_line}:AK{self.nex_line}",
            [
                [
                    {'numberFormat': {'type': 'DATE', 'pattern': 'dd.mm.yyyy'}},
                    {'numberFormat': {}},
                    {'numberFormat': {}},
                    {'numberFormat': {}},
                    {'numberFormat': {}},
                    {'numberFormat': {'type': 'CURRENCY', 'pattern': '#,##0.00[$ ₽]'}},
                    {'numberFormat': {'type': 'CURRENCY', 'pattern': '#,##0.00[$ ₽]'}},
                    {'numberFormat': {'type': 'CURRENCY', 'pattern': '#,##0.00[$ ₽]'}},
                    {'numberFormat': {}},
                    {'numberFormat': {'type': 'CURRENCY', 'pattern': '#,##0.00[$ ₽]'}},
                    {'numberFormat': {'type': 'CURRENCY', 'pattern': '#,##0.00[$ ₽]'}},
                    {'numberFormat': {'type': 'CURRENCY', 'pattern': '#,##0.00[$ ₽]'}},
                    {'numberFormat': {}},
                    {'numberFormat': {'type': 'CURRENCY', 'pattern': '#,##0.00[$ ₽]'}},
                    {'numberFormat': {'type': 'CURRENCY', 'pattern': '#,##0.00[$ ₽]'}},
                    {'numberFormat': {}},
                    {'numberFormat': {'type': 'CURRENCY', 'pattern': '#,##0.00[$ ₽]'}},
                    {'numberFormat': {'type': 'CURRENCY', 'pattern': '#,##0.00[$ ₽]'}},
                    {'numberFormat': {}},
                    {'numberFormat': {'type': 'CURRENCY', 'pattern': '#,##0.00[$ ₽]'}},
                    {'numberFormat': {'type': 'CURRENCY', 'pattern': '#,##0.00[$ ₽]'}},
                    {'numberFormat': {}},
                    {'numberFormat': {'type': 'CURRENCY', 'pattern': '#,##0.00[$ ₽]'}},
                    {'numberFormat': {'type': 'CURRENCY', 'pattern': '#,##0.00[$ ₽]'}},
                    {'numberFormat': {}},
                    {'numberFormat': {'type': 'CURRENCY', 'pattern': '#,##0.00[$ ₽]'}},
                    {'numberFormat': {'type': 'CURRENCY', 'pattern': '#,##0.00[$ ₽]'}},
                    {'numberFormat': {}},
                    {'numberFormat': {'type': 'CURRENCY', 'pattern': '#,##0.00[$ ₽]'}},
                    {'numberFormat': {'type': 'CURRENCY', 'pattern': '#,##0.00[$ ₽]'}},
                    {'numberFormat': {}},
                    {'numberFormat': {'type': 'CURRENCY', 'pattern': '#,##0.00[$ ₽]'}},
                    {'numberFormat': {}},
                    {'numberFormat': {'type': 'CURRENCY', 'pattern': '#,##0.00[$ ₽]'}},
                    {'numberFormat': {'type': 'CURRENCY', 'pattern': '#,##0.00[$ ₽]'}},
                    {'numberFormat': {'type': 'CURRENCY', 'pattern': '#,##0.00[$ ₽]'}},
                    {'numberFormat': {'type': 'CURRENCY', 'pattern': '#,##0.00[$ ₽]'}},
                ]
            ]
        )
        # Цвет фона ячеек
        if self.nex_line % 2 != 0:
            ss.prepare_setCellsFormat(f"A{self.nex_line}:AK{self.nex_line}",
                                      {"backgroundColor": functions.htmlColorToJSON("#fef8e3")},
                                      fields="userEnteredFormat.backgroundColor")

        # Бордер
        for j in range(self.sheet_width):
            ss.requests.append({"updateBorders": {
                "range": {"sheetId": ss.sheetId, "startRowIndex": self.nex_line - 1, "endRowIndex": self.nex_line,
                          "startColumnIndex": j,
                          "endColumnIndex": j + 1},
                "top": {"style": "SOLID", "width": 1, "color": {"red": 0, "green": 0, "blue": 0}}}})
            ss.requests.append({"updateBorders": {
                "range": {"sheetId": ss.sheetId, "startRowIndex": self.nex_line - 1, "endRowIndex": self.nex_line,
                          "startColumnIndex": j,
                          "endColumnIndex": j + 1},
                "right": {"style": "SOLID", "width": 1,
                          "color": {"red": 0, "green": 0, "blue": 0, "alpha": 1.0}}}})
            ss.requests.append({"updateBorders": {
                "range": {"sheetId": ss.sheetId, "startRowIndex": self.nex_line - 1, "endRowIndex": self.nex_line,
                          "startColumnIndex": j,
                          "endColumnIndex": j + 1},
                "left": {"style": "SOLID", "width": 1,
                         "color": {"red": 0, "green": 0, "blue": 0, "alpha": 1.0}}}})
            ss.requests.append({"updateBorders": {
                "range": {"sheetId": ss.sheetId, "startRowIndex": self.nex_line - 1, "endRowIndex": self.nex_line,
                          "startColumnIndex": j,
                          "endColumnIndex": j + 1},
                "bottom": {"style": "SOLID", "width": 1,
                           "color": {"red": 0, "green": 0, "blue": 0, "alpha": 1.0}}}})
        ss.runPrepared()

        # ------------------------------------------- Заполнение ИТОГО --------------------------------------
        # Вычисление последней строки в таблице
        for i, line_table in enumerate(self.spreadsheet['sheets'][0]['data'][0]['rowData']):
            try:
                if line_table['values'][0]['formattedValue'] == "ИТОГО":
                    # Если строка переписывается - итого на 1 поз вниз, если новая - на 2 поз
                    height_table = i + self.reprint
                    break
                else:
                    height_table = 4
            except KeyError:
                pass

        ss.prepare_setValues(f"A{height_table}:AK{height_table}",
                             [[f'ИТОГО',
                               "",
                               f"=SUM(C3:C{height_table - 1})",
                               f"=SUM(D3:D{height_table - 1})",
                               f"=SUM(E3:E{height_table - 1})",
                               f"=SUM(F3:F{height_table - 1})",
                               f"=SUM(G3:G{height_table - 1})",
                               f"=SUM(H3:H{height_table - 1})",
                               f"=SUM(I3:I{height_table - 1})",
                               f"=SUM(J3:J{height_table - 1})",
                               f"=IFERROR(ROUND(J{height_table}/I{height_table};2);0)",
                               f"=SUM(L3:L{height_table - 1})",
                               f"=SUM(M3:M{height_table - 1})",
                               f"=SUM(N3:N{height_table - 1})",
                               f"=IFERROR(ROUND(N{height_table}/M{height_table};2);0)",
                               f"=SUM(P3:P{height_table - 1})",
                               f"=SUM(Q3:Q{height_table - 1})",
                               f"=IFERROR(ROUND(Q{height_table}/P{height_table};2);0)",
                               f"=SUM(S3:S{height_table - 1})",
                               f"=SUM(T3:T{height_table - 1})",
                               f"=IFERROR(ROUND(T{height_table}/S{height_table};2);0)",
                               f"=SUM(V3:V{height_table - 1})",
                               f"=SUM(W3:W{height_table - 1})",
                               f"=IFERROR(ROUND(W{height_table}/V{height_table};2);0)",
                               f"=SUM(Y3:Y{height_table - 1})",
                               f"=SUM(Z3:Z{height_table - 1})",
                               f"=IFERROR(ROUND(Z{height_table}/Y{height_table};2);0)",
                               f"=SUM(AB3:AB{height_table - 1})",
                               f"=SUM(AC3:AC{height_table - 1})",
                               f"=IFERROR(ROUND(AC{height_table}/AB{height_table};2);0)",
                               f"=SUM(AE3:AE{height_table - 1})",
                               f"=SUM(AF3:AF{height_table - 1})",
                               f"=SUM(AG3:AG{height_table - 1})",
                               f"=SUM(AH3:AH{height_table - 1})",
                               f"=IFERROR(ROUND(AH{height_table}/AG{height_table};2);0)",
                               f"=SUM(AJ3:AJ{height_table - 1})",
                               f"=SUM(AK3:AK{height_table - 1})",
                               ]],
                             "ROWS")
        # Задание форматы вывода строки
        ss.prepare_setCellsFormats(
            f"A{height_table}:AK{height_table}",
            [
                [
                    {'numberFormat': {}},
                    {'numberFormat': {}},
                    {'numberFormat': {}},
                    {'numberFormat': {}},
                    {'numberFormat': {}},
                    {'numberFormat': {'type': 'CURRENCY', 'pattern': '#,##0.00[$ ₽]'}},
                    {'numberFormat': {'type': 'CURRENCY', 'pattern': '#,##0.00[$ ₽]'}},
                    {'numberFormat': {'type': 'CURRENCY', 'pattern': '#,##0.00[$ ₽]'}},
                    {'numberFormat': {}},
                    {'numberFormat': {'type': 'CURRENCY', 'pattern': '#,##0.00[$ ₽]'}},
                    {'numberFormat': {'type': 'CURRENCY', 'pattern': '#,##0.00[$ ₽]'}},
                    {'numberFormat': {'type': 'CURRENCY', 'pattern': '#,##0.00[$ ₽]'}},
                    {'numberFormat': {}},
                    {'numberFormat': {'type': 'CURRENCY', 'pattern': '#,##0.00[$ ₽]'}},
                    {'numberFormat': {'type': 'CURRENCY', 'pattern': '#,##0.00[$ ₽]'}},
                    {'numberFormat': {}},
                    {'numberFormat': {'type': 'CURRENCY', 'pattern': '#,##0.00[$ ₽]'}},
                    {'numberFormat': {'type': 'CURRENCY', 'pattern': '#,##0.00[$ ₽]'}},
                    {'numberFormat': {}},
                    {'numberFormat': {'type': 'CURRENCY', 'pattern': '#,##0.00[$ ₽]'}},
                    {'numberFormat': {'type': 'CURRENCY', 'pattern': '#,##0.00[$ ₽]'}},
                    {'numberFormat': {}},
                    {'numberFormat': {'type': 'CURRENCY', 'pattern': '#,##0.00[$ ₽]'}},
                    {'numberFormat': {'type': 'CURRENCY', 'pattern': '#,##0.00[$ ₽]'}},
                    {'numberFormat': {}},
                    {'numberFormat': {'type': 'CURRENCY', 'pattern': '#,##0.00[$ ₽]'}},
                    {'numberFormat': {'type': 'CURRENCY', 'pattern': '#,##0.00[$ ₽]'}},
                    {'numberFormat': {}},
                    {'numberFormat': {'type': 'CURRENCY', 'pattern': '#,##0.00[$ ₽]'}},
                    {'numberFormat': {'type': 'CURRENCY', 'pattern': '#,##0.00[$ ₽]'}},
                    {'numberFormat': {}},
                    {'numberFormat': {'type': 'CURRENCY', 'pattern': '#,##0.00[$ ₽]'}},
                    {'numberFormat': {}},
                    {'numberFormat': {'type': 'CURRENCY', 'pattern': '#,##0.00[$ ₽]'}},
                    {'numberFormat': {'type': 'CURRENCY', 'pattern': '#,##0.00[$ ₽]'}},
                    {'numberFormat': {'type': 'CURRENCY', 'pattern': '#,##0.00[$ ₽]'}},
                    {'numberFormat': {'type': 'CURRENCY', 'pattern': '#,##0.00[$ ₽]'}},
                ]
            ]
        )
        ss.prepare_setCellsFormat(f"A{height_table}:AK{height_table}",
                                  {'horizontalAlignment': 'RIGHT', 'textFormat': {'bold': True}})

        # Цвет фона ячеек
        ss.prepare_setCellsFormat(f"A{height_table}:AK{height_table}",
                                  {"backgroundColor": functions.htmlColorToJSON("#fce8b2")},
                                  fields="userEnteredFormat.backgroundColor")

        # Бордер
        for j in range(self.sheet_width):
            ss.requests.append({"updateBorders": {
                "range": {"sheetId": ss.sheetId, "startRowIndex": height_table - 1, "endRowIndex": height_table,
                          "startColumnIndex": j,
                          "endColumnIndex": j + 1},
                "top": {"style": "SOLID", "width": 1, "color": {"red": 0, "green": 0, "blue": 0}}}})
            ss.requests.append({"updateBorders": {
                "range": {"sheetId": ss.sheetId, "startRowIndex": height_table - 1, "endRowIndex": height_table,
                          "startColumnIndex": j,
                          "endColumnIndex": j + 1},
                "right": {"style": "SOLID", "width": 1,
                          "color": {"red": 0, "green": 0, "blue": 0, "alpha": 1.0}}}})
            ss.requests.append({"updateBorders": {
                "range": {"sheetId": ss.sheetId, "startRowIndex": height_table - 1, "endRowIndex": height_table,
                          "startColumnIndex": j,
                          "endColumnIndex": j + 1},
                "left": {"style": "SOLID", "width": 1,
                         "color": {"red": 0, "green": 0, "blue": 0, "alpha": 1.0}}}})
            ss.requests.append({"updateBorders": {
                "range": {"sheetId": ss.sheetId, "startRowIndex": height_table - 1, "endRowIndex": height_table,
                          "startColumnIndex": j,
                          "endColumnIndex": j + 1},
                "bottom": {"style": "SOLID", "width": 1,
                           "color": {"red": 0, "green": 0, "blue": 0, "alpha": 1.0}}}})

        ss.runPrepared()

    def open_googlesheet(self):
        """
        Открывает браузер с текущей гугл-таблицей
        """
        if not self.open_browser:
            logging.info(f'{__name__}: {str(datetime.now())[:-7]}:    Открытие файла-отчета в браузере...')
            self.show_dialog_variant(f'Открыть Google-отчет?',
                                     'Открыть Google-отчет?',
                                     webbrowser.open,
                                     self.spreadsheet['spreadsheetUrl']
                                     )
            self.open_browser = True

    def sms_report(self):
        """
        Составляет текстовую версию финансового отчета
        :return: str
        """
        logging.info(f'{__name__}: {str(datetime.now())[:-7]}:    Составление SMS-отчета...')
        resporse = 'Отчет по аквапарку за '
        if self.finreport_dict['Дата'][0] == self.finreport_dict['Дата'][1] - timedelta(1):
            resporse += f'{datetime.strftime(self.finreport_dict["Дата"][0], "%d.%m.%Y")}:\n'
        else:
            resporse += f'{datetime.strftime(self.finreport_dict["Дата"][0], "%d.%m.%Y")} - {datetime.strftime(self.finreport_dict["Дата"][1] - timedelta(1), "%d.%m.%Y")}:\n'
        if self.finreport_dict['ИТОГО'][1]:
            resporse += f'Люди - {self.finreport_dict["Кол-во проходов"][0]};\n'
            resporse += f'По аквапарку - {self.finreport_dict["Билеты аквапарка"][1] + self.finreport_dict["Билеты аквапарка КОРП"][1]:.2f} ₽;\n'
            resporse += f'По общепиту - {self.finreport_dict["Общепит"][1]:.2f} ₽;\n'
            resporse += f'Термозона - {self.finreport_dict["Термозона"][1] + self.finreport_dict["Термозона КОРП"][1]:.2f} ₽;\n'
            resporse += f'Прочее - {self.finreport_dict["Прочее"][1]:.2f} ₽;\n'
            resporse += f'Общая по БАРСу - {self.finreport_dict["ИТОГО"][1]:.2f} ₽;\n'
            resporse += f'ONLINE продажи - {self.finreport_dict["Online Продажи"][1]:.2f} ₽;\n'
        if self.itog_report_org2['Итого по отчету'][1]:
            # resporse += 'Отчет по пляжу за '
            # if beach_report['Дата'][0] + timedelta(1) == beach_report['Дата'][1]:
            #     resporse += f'{datetime.strftime(beach_report["Дата"][0], "%d.%m.%Y")}:\n'
            # else:
            #     resporse += f'{datetime.strftime(beach_report["Дата"][0], "%d.%m.%Y")} - {datetime.strftime(beach_report["Дата"][0], "%d.%m.%Y")}:\n'
            resporse += f'Люди (пляж) - {self.itog_report_org2["Летняя зона | БЕЗЛИМИТ | 1 проход"][0]};\n'
            resporse += f'Итого по пляжу - {self.itog_report_org2["Итого по отчету"][1]:.2f} ₽;\n'
        resporse += f'Без ЧП.'
        return resporse

    def send_message_to_telegram(self):
        """
        Отправка отчета в telegram
        """
        logging.info(f'{__name__}: {str(datetime.now())[:-7]}:    Отправка SMS-отчета в Telegram-канал...')
        if self.telegram_proxy_use:
            logging.info(f'{__name__}: {str(datetime.now())[:-7]}:    Соединение с прокси-сервером {self.telegram_proxy_ip}...')
            if self.telegram_proxy_auth:
                socks.setdefaultproxy(proxy_type=getattr(socks, self.telegram_proxy_type),
                                      addr=self.telegram_proxy_ip,
                                      port=int(self.telegram_proxy_port),
                                      rdns=True,
                                      username=self.telegram_proxy_username,
                                      password=self.telegram_proxy_password,
                                      )
            else:
                socks.setdefaultproxy(proxy_type=getattr(socks, self.telegram_proxy_type),
                                      addr=self.telegram_proxy_ip,
                                      port=int(self.telegram_proxy_port),
                                      )
            socket.socket = socks.socksocket
        bot = telepot.Bot(self.telegram_token)
        for line in self.sms_report_list:
            logging.info(f'{__name__}: {str(datetime.now())[:-7]}:    '
                         f'Отправка сообщения {self.sms_report_list.index(line) + 1} из {len(self.sms_report_list)}')
            bot.sendMessage(self.telegram_chanel_id, line)
        # if self.telegram_proxy_use:
        #     logging.info(
        #         f'{__name__}: {str(datetime.now())[:-7]}:    Разъединение с прокси-сервером {self.telegram_proxy_ip}...')
        #     socks.setdefaultproxy()
        #     socket.socket = socks.socksocket

    def save_organisation_total(self, itog_report):
        """
        Сохраняет Итоговый отчет в Excel
        """
        organisation_total = {}
        for key in itog_report:
            organisation_total[itog_report[key][3]] = {}
        for key in itog_report:
            organisation_total[itog_report[key][3]][itog_report[key][2]] = []
        for key in itog_report:
            organisation_total[itog_report[key][3]][itog_report[key][2]].append((key, itog_report[key][0], itog_report[key][1]))

        # определяем стили
        h1 = Font(name='Times New Roman',
                  size=18,
                  bold=True,
                  italic=False,
                  vertAlign=None,
                  underline='none',
                  strike=False,
                  color='FF000000')
        h2 = Font(name='Times New Roman',
                  size=14,
                  bold=True,
                  italic=False,
                  vertAlign=None,
                  underline='none',
                  strike=False,
                  color='FF000000')
        h3 = Font(name='Times New Roman',
                  size=11,
                  bold=True,
                  italic=False,
                  vertAlign=None,
                  underline='none',
                  strike=False,
                  color='FF000000')
        font = Font(name='Times New Roman',
                    size=9,
                    bold=False,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='FF000000')
        font_bold = Font(name='Times New Roman',
                    size=9,
                    bold=True,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='FF000000')

        fill = PatternFill(fill_type='solid',
                           start_color='c1c1c1',
                           end_color='c2c2c2')
        align_top = Alignment(horizontal='general',
                              vertical='top',
                              text_rotation=0,
                              wrap_text=False,
                              shrink_to_fit=False,
                              indent=0,
                              )
        align_bottom = Alignment(horizontal='general',
                              vertical='bottom',
                              text_rotation=0,
                              wrap_text=False,
                              shrink_to_fit=False,
                              indent=0,
                              )

        border = Border(left=Side(border_style='thin',
                                  color='FF000000'),
                        right=Side(border_style='thin',
                                   color='FF000000'),
                        top=Side(border_style='thin',
                                 color='FF000000'),
                        bottom=Side(border_style='thin',
                                    color='FF000000'),
                        diagonal=Side(border_style='thin',
                                      color='FF000000'),
                        diagonal_direction=0,
                        outline=Side(border_style='thin',
                                     color='FF000000'),
                        vertical=Side(border_style='thin',
                                      color='FF000000'),
                        horizontal=Side(border_style='thin',
                                        color='FF000000')
                        )
        border_top_bottom = Border(bottom=Side(border_style='thin', color='FF000000'),
                                   top=Side(border_style='thin', color='FF000000'),
                                   )
        border_right = Border(right=Side(border_style='thin', color='FF000000'))
        border_left = Border(left=Side(border_style='thin', color='FF000000'))
        border_top = Border(top=Side(border_style='thin', color='FF000000'))
        border_left_top = Border(top=Side(border_style='thin', color='FF000000'),
                                 left=Side(border_style='thin', color='FF000000'),
                                 )
        border_right_top = Border(top=Side(border_style='thin', color='FF000000'),
                                  right=Side(border_style='thin', color='FF000000'),
                                  )
        align_center = Alignment(horizontal='center',
                                 vertical='bottom',
                                 text_rotation=0,
                                 wrap_text=False,
                                 shrink_to_fit=False,
                                 indent=0)
        align_left = Alignment(horizontal='left',
                               vertical='bottom',
                               text_rotation=0,
                               wrap_text=False,
                               shrink_to_fit=False,
                               indent=0)
        number_format = 'General'
        protection = Protection(locked=True,
                                hidden=False)

        column = ['', 'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']

        self.row = '0'
        def next_row():
            self.row = str(int(self.row) + 1)
            return self.row

        # объект
        wb = Workbook()

        # активный лист
        ws = wb.active

        # название страницы
        # ws = wb.create_sheet('первая страница', 0)
        ws.title = 'Итоговый отчет'
        # шрифты
        ws['C1'].font = h1
        # выравнивание
        ws['C1'].alignment = align_left

        # Ширина стролбцов
        ws.column_dimensions['A'].width = 1 / 7 * 8
        ws.column_dimensions['B'].width = 1 / 7 * 8
        ws.column_dimensions['C'].width = 1 / 7 * 80
        ws.column_dimensions['D'].width = 1 / 7 * 8
        ws.column_dimensions['E'].width = 1 / 7 * 88
        ws.column_dimensions['F'].width = 1 / 7 * 8
        ws.column_dimensions['G'].width = 1 / 7 * 24
        ws.column_dimensions['H'].width = 1 / 7 * 8
        ws.column_dimensions['I'].width = 1 / 7 * 80
        ws.column_dimensions['J'].width = 1 / 7 * 8
        ws.column_dimensions['K'].width = 1 / 7 * 144
        ws.column_dimensions['L'].width = 1 / 7 * 144
        ws.column_dimensions['M'].width = 1 / 7 * 8

        # значение ячейки
        # ws['A1'] = "Hello!"

        ws[column[3] + next_row()] = 'Итоговый отчет'
        ws.merge_cells(start_row=self.row, start_column=3, end_row=self.row, end_column=12)
        ws[column[1] + next_row()] = ''
        ws[column[3] + next_row()] = organisation_total['Организация']['Организация'][0][0]
        ws.merge_cells(start_row=self.row, start_column=3, end_row=self.row, end_column=12)
        ws[column[3] + self.row].font = font
        ws[column[3] + self.row].alignment = align_top
        ws[column[1] + next_row()] = ''

        ws[column[3] + next_row()] = 'За период с:'
        ws[column[3] + self.row].font = font
        ws[column[3] + self.row].alignment = align_top
        ws[column[5] + self.row] = itog_report['Дата'][0].strftime("%d.%m.%Y")
        ws[column[5] + self.row].font = font_bold
        ws[column[5] + self.row].alignment = align_top
        ws[column[7] + self.row] = 'По:'
        ws[column[7] + self.row].font = font
        ws[column[7] + self.row].alignment = align_top
        ws[column[9] + self.row] = (itog_report['Дата'][1] - timedelta(1)).strftime("%d.%m.%Y")
        ws[column[9] + self.row].font = font_bold
        ws[column[9] + self.row].alignment = align_top

        # ТАБЛИЦА
        def merge_table():
            ws.merge_cells(start_row=self.row, start_column=2, end_row=self.row, end_column=9)
            ws.merge_cells(start_row=self.row, start_column=10, end_row=self.row, end_column=11)
            ws.merge_cells(start_row=self.row, start_column=12, end_row=self.row, end_column=13)
            ws[column[2] + self.row].font = font
            ws[column[10] + self.row].font = font
            ws[column[12] + self.row].font = font
            ws[column[2] + self.row].alignment = align_top
            ws[column[10] + self.row].alignment = align_top
            ws[column[12] + self.row].alignment = align_top
            b = 2
            while b <= 13:
                ws[column[b] + self.row].border = border
                b += 1

        def merge_table_h3():
            ws.merge_cells(start_row=self.row, start_column=2, end_row=self.row, end_column=9)
            ws.merge_cells(start_row=self.row, start_column=10, end_row=self.row, end_column=11)
            ws.merge_cells(start_row=self.row, start_column=12, end_row=self.row, end_column=13)
            ws[column[2] + self.row].font = h3
            ws[column[10] + self.row].font = h3
            ws[column[12] + self.row].font = h3
            ws[column[2] + self.row].alignment = align_top
            ws[column[10] + self.row].alignment = align_top
            ws[column[12] + self.row].alignment = align_top
            ws[column[2] + self.row].border = border_left
            ws[column[13] + self.row].border = border_right

        def merge_table_h2():
            ws.merge_cells(start_row=self.row, start_column=2, end_row=self.row, end_column=9)
            ws.merge_cells(start_row=self.row, start_column=10, end_row=self.row, end_column=11)
            ws.merge_cells(start_row=self.row, start_column=12, end_row=self.row, end_column=13)
            ws[column[2] + self.row].font = h2
            ws[column[10] + self.row].font = h2
            ws[column[12] + self.row].font = h2
            ws[column[2] + self.row].alignment = align_top
            ws[column[10] + self.row].alignment = align_top
            ws[column[12] + self.row].alignment = align_top
            ws[column[2] + self.row].border = border_left
            ws[column[13] + self.row].border = border_right

        def merge_width_h2():
            ws.merge_cells(start_row=self.row, start_column=2, end_row=self.row, end_column=13)
            ws[column[2] + self.row].font = h2
            ws[column[2] + self.row].alignment = align_top
            b = 2
            while b <= 13:
                if b == 2:
                    ws[column[b] + self.row].border = border_left_top
                elif b == 13:
                    ws[column[b] + self.row].border = border_right_top
                else:
                    ws[column[b] + self.row].border = border_top
                b += 1

        def merge_width_h3():
            ws.merge_cells(start_row=self.row, start_column=2, end_row=self.row, end_column=13)
            ws[column[2] + self.row].font = h3
            ws[column[2] + self.row].alignment = align_top
            b = 2
            while b <= 13:
                if b == 2:
                    ws[column[b] + self.row].border = border_left
                elif b == 13:
                    ws[column[b] + self.row].border = border_right
                b += 1

        ws[column[2] + next_row()] = 'Название'
        ws[column[10] + self.row] = 'Количество'
        ws[column[12] + self.row] = 'Сумма'
        merge_table()
        ws[column[2] + self.row].font = h3
        ws[column[10] + self.row].font = h3
        ws[column[12] + self.row].font = h3
        ws[column[2] + self.row].alignment = align_top
        ws[column[10] + self.row].alignment = align_top
        ws[column[12] + self.row].alignment = align_top

        groups = [
            'Депозит',
            'Карты',
            'Услуги',
            'Товары',
            'Платные зоны',
        ]
        all_count = 0
        all_sum = 0
        try:
            for gr in groups:
                ws[column[2] + next_row()] = gr
                merge_width_h2()
                group_count = 0
                group_sum = 0
                for group in organisation_total[gr]:
                    ws[column[2] + next_row()] = group
                    merge_width_h3()
                    service_count = 0
                    service_sum = 0
                    for service in organisation_total[gr][group]:
                        try:
                            service_count += service[1]
                            service_sum += service[2]
                        except TypeError:
                            pass
                        ws[column[2] + next_row()] = service[0]
                        ws[column[10] + self.row] = service[1]
                        ws[column[12] + self.row] = service[2]
                        ws[column[12] + self.row].number_format = '#,##0.00 ₽'
                        merge_table()
                    ws[column[10] + next_row()] = service_count
                    ws[column[12] + self.row] = service_sum
                    ws[column[12] + self.row].number_format = '#,##0.00 ₽'
                    merge_table_h3()
                    group_count += service_count
                    group_sum += service_sum
                ws[column[10] + next_row()] = group_count
                ws[column[12] + self.row] = group_sum
                ws[column[12] + self.row].number_format = '#,##0.00 ₽'
                merge_table_h2()
                all_count += group_count
                all_sum += group_sum
                group_count = 0
                group_sum = 0
        except KeyError:
            pass

        if all_sum == organisation_total["Итого по отчету"][""][0][2]:
            ws[column[2] + next_row()] = organisation_total['Итого по отчету'][''][0][0]
            ws[column[10] + self.row] = organisation_total['Итого по отчету'][''][0][1]
            ws[column[12] + self.row] = organisation_total["Итого по отчету"][""][0][2]
            self.total_report_sum = all_sum
        else:
            logging.error(f'{__name__}: {str(datetime.now())[:-7]}:    Ошибка: Итоговые суммы не совпадают. "Итого по отчету" '
                          f'из Барса не совпадает с итоговой суммой по формируемым строкам.')
            self.show_dialog(f'Ошибка: Итоговые суммы не совпадают',
                             '"Итого по отчету" из Барса не совпадает с итоговой суммой по формируемым строкам.'
                             )
            return None
        ws[column[12] + self.row].number_format = '#,##0.00 ₽'
        merge_table_h2()
        ws[column[2] + self.row].alignment = align_bottom
        ws[column[10] + self.row].alignment = align_bottom
        ws[column[12] + self.row].alignment = align_bottom
        b = 2
        while b <= 13:
            ws[column[b] + self.row].border = border_top_bottom
            b += 1
        end_line = int(self.row)

        # раскрвшивание фона для заголовков
        i = 2
        while i <= 13:
            ws[column[i] + '6'].fill = fill
            i += 1

        # обводка
        # ws['A3'].border = border

        # вручную устанавливаем высоту первой строки
        rd = ws.row_dimensions[1]
        rd.height = 21.75

        # увеличиваем все строки по высоте
        max_row = ws.max_row
        i = 2
        while i <= max_row:
            rd = ws.row_dimensions[i]
            rd.height = 18
            i += 1

        # Высота строк
        ws.row_dimensions[2].height = 5.25
        ws.row_dimensions[4].height = 6.75
        ws.row_dimensions[end_line].height = 30.75

        # выравнивание столбца
        for cellObj in ws['A2:A5']:
            for cell in cellObj:
                ws[cell.coordinate].alignment = align_left

        # перетягивание ячеек
        # https://stackoverflow.com/questions/13197574/openpyxl-adjust-column-width-size
        # dims = {}
        # for row in ws.rows:
        #     for cell in row:
        #         if cell.value:
        #             dims[cell.column] = max((dims.get(cell.column, 0), len(cell.value)))
        # for col, value in dims.items():
        #     # value * коэфициент
        #     ws.column_dimensions[col].width = value * 1.5

        # сохранение файла в текущую директорию
        if itog_report['Дата'][0] == itog_report["Дата"][1] - timedelta(1):
            date_ = datetime.strftime(itog_report["Дата"][0], "%Y-%m-%d")
        else:
            date_ = f'{datetime.strftime(itog_report["Дата"][0], "%Y-%m-%d")} - ' \
                    f'{datetime.strftime(itog_report["Дата"][1] - timedelta(1), "%Y-%m-%d")}'
        path = self.local_folder + self.path + date_ + \
               f' Итоговый отчет по {organisation_total["Организация"]["Организация"][0][0]}' + ".xlsx"
        logging.info(f'{__name__}: {str(datetime.now())[:-7]}:    Сохранение Итогового отчета '
                     f'по {organisation_total["Организация"]["Организация"][0][0]} в {path}')
        path = self.create_path(path)
        self.save_file(path, wb)
        return path

    def save_cashdesk_report(self, cashdesk_report):
        """
        Сохраняет Суммовой отчет в Excel
        """
        # определяем стили
        h1 = Font(name='Times New Roman',
                  size=18,
                  bold=True,
                  italic=False,
                  vertAlign=None,
                  underline='none',
                  strike=False,
                  color='FF000000')
        h2 = Font(name='Times New Roman',
                  size=14,
                  bold=True,
                  italic=False,
                  vertAlign=None,
                  underline='none',
                  strike=False,
                  color='FF000000')
        h3 = Font(name='Times New Roman',
                  size=10,
                  bold=True,
                  italic=False,
                  vertAlign=None,
                  underline='none',
                  strike=False,
                  color='FF000000')
        font = Font(name='Times New Roman',
                    size=9,
                    bold=False,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='FF000000')
        font_bold = Font(name='Times New Roman',
                    size=9,
                    bold=True,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='FF000000')
        font_red = Font(name='Times New Roman',
                         size=9,
                         bold=True,
                         italic=False,
                         vertAlign=None,
                         underline='none',
                         strike=False,
                         color='FFFF0000')

        fill = PatternFill(fill_type='solid',
                           start_color='c1c1c1',
                           end_color='c2c2c2')
        align_top = Alignment(horizontal='general',
                              vertical='top',
                              text_rotation=0,
                              wrap_text=False,
                              shrink_to_fit=False,
                              indent=0,
                              )
        align_bottom = Alignment(horizontal='general',
                              vertical='bottom',
                              text_rotation=0,
                              wrap_text=False,
                              shrink_to_fit=False,
                              indent=0,
                              )

        border = Border(left=Side(border_style='thin',
                                  color='FF000000'),
                        right=Side(border_style='thin',
                                   color='FF000000'),
                        top=Side(border_style='thin',
                                 color='FF000000'),
                        bottom=Side(border_style='thin',
                                    color='FF000000'),
                        diagonal=Side(border_style='thin',
                                      color='FF000000'),
                        diagonal_direction=0,
                        outline=Side(border_style='thin',
                                     color='FF000000'),
                        vertical=Side(border_style='thin',
                                      color='FF000000'),
                        horizontal=Side(border_style='thin',
                                        color='FF000000')
                        )
        border_top_bottom = Border(bottom=Side(border_style='thin', color='FF000000'),
                                   top=Side(border_style='thin', color='FF000000'),
                                   )
        border_right = Border(right=Side(border_style='thin', color='FF000000'))
        border_left = Border(left=Side(border_style='thin', color='FF000000'))
        border_top = Border(top=Side(border_style='thin', color='FF000000'))
        border_left_top = Border(top=Side(border_style='thin', color='FF000000'),
                                 left=Side(border_style='thin', color='FF000000'),
                                 )
        border_right_top = Border(top=Side(border_style='thin', color='FF000000'),
                                  right=Side(border_style='thin', color='FF000000'),
                                  )
        align_center = Alignment(horizontal='center',
                                 vertical='bottom',
                                 text_rotation=0,
                                 wrap_text=False,
                                 shrink_to_fit=False,
                                 indent=0)
        align_left = Alignment(horizontal='left',
                               vertical='bottom',
                               text_rotation=0,
                               wrap_text=False,
                               shrink_to_fit=False,
                               indent=0)
        number_format = 'General'
        protection = Protection(locked=True,
                                hidden=False)

        column = ['', 'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']

        self.row = '0'
        def next_row():
            self.row = str(int(self.row) + 1)
            return self.row

        # объект
        wb = Workbook()

        # активный лист
        ws = wb.active

        # название страницы
        # ws = wb.create_sheet('первая страница', 0)
        ws.title = 'Суммовой отчет по чековой ленте'
        # шрифты
        ws['A1'].font = h1
        # выравнивание
        ws['A1'].alignment = align_left

        # Ширина стролбцов
        ws.column_dimensions['A'].width = 1 / 7 * 124
        ws.column_dimensions['B'].width = 1 / 7 * 88
        ws.column_dimensions['C'].width = 1 / 7 * 28
        ws.column_dimensions['D'].width = 1 / 7 * 24
        ws.column_dimensions['E'].width = 1 / 7 * 32
        ws.column_dimensions['F'].width = 1 / 7 * 1
        ws.column_dimensions['G'].width = 1 / 7 * 79
        ws.column_dimensions['H'].width = 1 / 7 * 3
        ws.column_dimensions['I'].width = 1 / 7 * 5
        ws.column_dimensions['J'].width = 1 / 7 * 96
        ws.column_dimensions['K'].width = 1 / 7 * 88
        ws.column_dimensions['L'].width = 1 / 7 * 8
        ws.column_dimensions['M'].width = 1 / 7 * 80
        ws.column_dimensions['N'].width = 1 / 7 * 96

        # значение ячейки
        # ws['A1'] = "Hello!"

        ws[column[1] + next_row()] = 'Суммовой отчет по чековой ленте'
        ws.merge_cells(start_row=self.row, start_column=1, end_row=self.row, end_column=len(column)-1)
        # шрифты
        ws[column[1] + self.row].font = h1
        # выравнивание
        ws[column[1] + self.row].alignment = align_left
        # Высота строк
        ws.row_dimensions[1].height = 24

        ws[column[1] + next_row()] = f'{cashdesk_report["Организация"][0][0]}'
        ws.merge_cells(start_row=self.row, start_column=1, end_row=self.row, end_column=len(column) - 1)
        ws[column[1] + self.row].font = font
        ws[column[1] + self.row].alignment = align_top

        ws[column[1] + next_row()] = 'За период с:'
        ws[column[1] + self.row].font = font
        ws[column[1] + self.row].alignment = align_top
        ws[column[2] + self.row] = cashdesk_report['Дата'][0][0].strftime("%d.%m.%Y")
        ws.merge_cells(start_row=self.row, start_column=2, end_row=self.row, end_column=3)
        ws[column[2] + self.row].font = font_bold
        ws[column[2] + self.row].alignment = align_top
        ws[column[4] + self.row] = 'по'
        ws[column[4] + self.row].font = font
        ws[column[4] + self.row].alignment = align_top
        ws[column[5] + self.row] = (cashdesk_report['Дата'][0][1] - timedelta(1)).strftime("%d.%m.%Y")
        ws.merge_cells(start_row=self.row, start_column=5, end_row=self.row, end_column=7)
        ws[column[5] + self.row].font = font_bold
        ws[column[5] + self.row].alignment = align_top

        # ТАБЛИЦА
        def merge_table():
            ws.merge_cells(start_row=self.row, start_column=1, end_row=self.row, end_column=2)
            ws.merge_cells(start_row=self.row, start_column=3, end_row=self.row, end_column=6)
            ws.merge_cells(start_row=self.row, start_column=7, end_row=self.row, end_column=9)
            ws.merge_cells(start_row=self.row, start_column=11, end_row=self.row, end_column=12)
            ws[column[1] + self.row].font = font
            ws[column[3] + self.row].font = font
            ws[column[7] + self.row].font = font
            ws[column[10] + self.row].font = font
            ws[column[11] + self.row].font = font
            ws[column[13] + self.row].font = font
            ws[column[14] + self.row].font = font
            ws[column[1] + self.row].alignment = align_top
            ws[column[3] + self.row].alignment = align_top
            ws[column[7] + self.row].alignment = align_top
            ws[column[10] + self.row].alignment = align_top
            ws[column[11] + self.row].alignment = align_top
            ws[column[13] + self.row].alignment = align_top
            ws[column[14] + self.row].alignment = align_top
            ws[column[3] + self.row].number_format = '#,##0.00 ₽'
            ws[column[7] + self.row].number_format = '#,##0.00 ₽'
            ws[column[10] + self.row].number_format = '#,##0.00 ₽'
            ws[column[11] + self.row].number_format = '#,##0.00 ₽'
            ws[column[13] + self.row].number_format = '#,##0.00 ₽'
            ws[column[14] + self.row].number_format = '#,##0.00 ₽'
            b = 1
            while b < len(column):
                ws[column[b] + self.row].border = border
                b += 1

        def merge_table_h3():
            ws.merge_cells(start_row=self.row, start_column=1, end_row=self.row, end_column=2)
            ws.merge_cells(start_row=self.row, start_column=3, end_row=self.row, end_column=6)
            ws.merge_cells(start_row=self.row, start_column=7, end_row=self.row, end_column=9)
            ws.merge_cells(start_row=self.row, start_column=11, end_row=self.row, end_column=12)
            ws[column[1] + self.row].font = h3
            ws[column[3] + self.row].font = h3
            ws[column[7] + self.row].font = h3
            ws[column[10] + self.row].font = h3
            ws[column[11] + self.row].font = h3
            ws[column[13] + self.row].font = h3
            ws[column[14] + self.row].font = h3
            ws[column[1] + self.row].alignment = align_top
            ws[column[3] + self.row].alignment = align_top
            ws[column[7] + self.row].alignment = align_top
            ws[column[10] + self.row].alignment = align_top
            ws[column[11] + self.row].alignment = align_top
            ws[column[13] + self.row].alignment = align_top
            ws[column[14] + self.row].alignment = align_top
            ws[column[3] + self.row].number_format = '#,##0.00 ₽'
            ws[column[7] + self.row].number_format = '#,##0.00 ₽'
            ws[column[10] + self.row].number_format = '#,##0.00 ₽'
            ws[column[11] + self.row].number_format = '#,##0.00 ₽'
            ws[column[13] + self.row].number_format = '#,##0.00 ₽'
            ws[column[14] + self.row].number_format = '#,##0.00 ₽'
            b = 1
            while b < len(column):
                ws[column[b] + self.row].border = border
                b += 1

        def merge_table_red():
            ws.merge_cells(start_row=self.row, start_column=1, end_row=self.row, end_column=2)
            ws.merge_cells(start_row=self.row, start_column=3, end_row=self.row, end_column=6)
            ws.merge_cells(start_row=self.row, start_column=7, end_row=self.row, end_column=9)
            ws.merge_cells(start_row=self.row, start_column=11, end_row=self.row, end_column=12)
            ws[column[1] + self.row].font = font_red
            ws[column[3] + self.row].font = font_red
            ws[column[7] + self.row].font = font_red
            ws[column[10] + self.row].font = font_red
            ws[column[11] + self.row].font = font_red
            ws[column[13] + self.row].font = font_red
            ws[column[14] + self.row].font = font_red
            ws[column[1] + self.row].alignment = align_top
            ws[column[3] + self.row].alignment = align_top
            ws[column[7] + self.row].alignment = align_top
            ws[column[10] + self.row].alignment = align_top
            ws[column[11] + self.row].alignment = align_top
            ws[column[13] + self.row].alignment = align_top
            ws[column[14] + self.row].alignment = align_top
            ws[column[3] + self.row].number_format = '#,##0.00 ₽'
            ws[column[7] + self.row].number_format = '#,##0.00 ₽'
            ws[column[10] + self.row].number_format = '#,##0.00 ₽'
            ws[column[11] + self.row].number_format = '#,##0.00 ₽'
            ws[column[13] + self.row].number_format = '#,##0.00 ₽'
            ws[column[14] + self.row].number_format = '#,##0.00 ₽'
            b = 1
            while b < len(column):
                ws[column[b] + self.row].border = border
                b += 1

        def merge_width_red():
            ws.merge_cells(start_row=self.row, start_column=1, end_row=self.row, end_column=len(column) - 1)
            ws[column[1] + self.row].font = font_red
            ws[column[1] + self.row].alignment = align_top
            b = 1
            while b < len(column):
                if b == 1:
                    ws[column[b] + self.row].border = border_left
                elif b == len(column) - 1:
                    ws[column[b] + self.row].border = border_right
                else:
                    ws[column[b] + self.row].border = border
                b += 1

        ws[column[1] + next_row()] = 'Касса №'
        ws[column[3] + self.row] = 'Сумма'
        ws[column[7] + self.row] = 'Наличными'
        ws[column[10] + self.row] = 'Безналичными'
        ws[column[11] + self.row] = 'Со счета'
        ws[column[13] + self.row] = 'Бонусами'
        ws[column[14] + self.row] = 'LSI'
        merge_table_h3()
        # раскрвшивание фона для заголовков
        b = 1
        while b < len(column):
            ws[column[b] + self.row].fill = fill
            b += 1

        for typpe in cashdesk_report:
            if typpe != 'Дата' and typpe != 'Организация':
                if typpe != 'Итого':
                    ws[column[1] + next_row()] = typpe
                    merge_width_red()
                for line in cashdesk_report[typpe]:
                    ws[column[1] + next_row()] = line[0]
                    ws[column[3] + self.row] = line[1]
                    ws[column[7] + self.row] = line[2]
                    ws[column[10] + self.row] = line[3]
                    ws[column[11] + self.row] = line[4]
                    ws[column[13] + self.row] = line[5]
                    ws[column[14] + self.row] = line[6]-line[3]
                    if line[0] == 'Итого':
                        merge_table_red()
                    elif line[0] == 'Итого по отчету':
                        merge_table_h3()
                    else:
                        merge_table()

        # увеличиваем все строки по высоте
        max_row = ws.max_row
        i = 2
        while i <= max_row:
            rd = ws.row_dimensions[i]
            rd.height = 18
            i += 1
        if cashdesk_report['Дата'][0][0] == cashdesk_report["Дата"][0][1] - timedelta(1):
            date_ = datetime.strftime(cashdesk_report["Дата"][0][0], "%Y-%m-%d")
        else:
            date_ = f'{datetime.strftime(cashdesk_report["Дата"][0][0], "%Y-%m-%d")} - ' \
                    f'{datetime.strftime(cashdesk_report["Дата"][0][1] - timedelta(1), "%Y-%m-%d")}'
        path = self.local_folder + self.path + date_ + f' Суммовой отчет по {cashdesk_report["Организация"][0][0]}' + ".xlsx"
        logging.info(f'{__name__}: {str(datetime.now())[:-7]}:    Сохранение Суммового отчета '
                     f'по {cashdesk_report["Организация"][0][0]} в {path}')
        path = self.create_path(path)
        self.save_file(path, wb)
        return path

    def save_client_count_totals(self, client_count_totals_org):
        """
        Сохраняет отчет по количеству клиентов за день в Excel
        """
        # определяем стили
        h1 = Font(name='Times New Roman',
                  size=18,
                  bold=True,
                  italic=False,
                  vertAlign=None,
                  underline='none',
                  strike=False,
                  color='FF000000')
        font = Font(name='Times New Roman',
                    size=9,
                    bold=False,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='FF000000')
        font_bold = Font(name='Times New Roman',
                    size=9,
                    bold=True,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='FF000000')
        fill = PatternFill(fill_type='solid',
                           start_color='c1c1c1',
                           end_color='c2c2c2')
        align_top = Alignment(horizontal='general',
                              vertical='top',
                              text_rotation=0,
                              wrap_text=False,
                              shrink_to_fit=False,
                              indent=0,
                              )
        border = Border(left=Side(border_style='thin',
                                  color='FF000000'),
                        right=Side(border_style='thin',
                                   color='FF000000'),
                        top=Side(border_style='thin',
                                 color='FF000000'),
                        bottom=Side(border_style='thin',
                                    color='FF000000'),
                        diagonal=Side(border_style='thin',
                                      color='FF000000'),
                        diagonal_direction=0,
                        outline=Side(border_style='thin',
                                     color='FF000000'),
                        vertical=Side(border_style='thin',
                                      color='FF000000'),
                        horizontal=Side(border_style='thin',
                                        color='FF000000')
                        )
        align_left = Alignment(horizontal='left',
                               vertical='bottom',
                               text_rotation=0,
                               wrap_text=False,
                               shrink_to_fit=False,
                               indent=0)
        number_format = 'General'
        protection = Protection(locked=True,
                                hidden=False)

        column = ['', 'A', 'B', 'C', 'D', 'E']

        self.row = '0'
        def next_row():
            self.row = str(int(self.row) + 1)
            return self.row

        # объект
        wb = Workbook()

        # активный лист
        ws = wb.active

        # название страницы
        # ws = wb.create_sheet('первая страница', 0)
        ws.title = 'Количество человек за день'
        # шрифты
        ws['A1'].font = h1
        # выравнивание
        ws['A1'].alignment = align_left

        # Ширина стролбцов
        ws.column_dimensions['A'].width = 1 / 7 * 124
        ws.column_dimensions['B'].width = 1 / 7 * 21
        ws.column_dimensions['C'].width = 1 / 7 * 95
        ws.column_dimensions['D'].width = 1 / 7 * 24
        ws.column_dimensions['E'].width = 1 / 7 * 80

        # значение ячейки
        # ws['A1'] = "Hello!"

        ws[column[1] + next_row()] = 'Количество человек за день'
        ws.merge_cells(start_row=self.row, start_column=1, end_row=self.row, end_column=len(column)-1)
        # шрифты
        ws[column[1] + self.row].font = h1
        # выравнивание
        ws[column[1] + self.row].alignment = align_left
        # Высота строк
        ws.row_dimensions[1].height = 24

        ws[column[1] + next_row()] = f'{client_count_totals_org[0][0]}'
        ws.merge_cells(start_row=self.row, start_column=1, end_row=self.row, end_column=len(column) - 1)
        ws[column[1] + self.row].font = font
        ws[column[1] + self.row].alignment = align_top

        ws[column[1] + next_row()] = 'За период с:'
        ws[column[1] + self.row].font = font
        ws[column[1] + self.row].alignment = align_top
        ws[column[2] + self.row] = client_count_totals_org[1][0].strftime("%d.%m.%Y")
        ws.merge_cells(start_row=self.row, start_column=2, end_row=self.row, end_column=3)
        ws[column[2] + self.row].font = font_bold
        ws[column[2] + self.row].alignment = align_top
        ws[column[4] + self.row] = 'по'
        ws[column[4] + self.row].font = font
        ws[column[4] + self.row].alignment = align_top
        ws[column[5] + self.row] = (client_count_totals_org[-2][0]).strftime("%d.%m.%Y")
        ws.merge_cells(start_row=self.row, start_column=5, end_row=self.row, end_column=7)
        ws[column[5] + self.row].font = font_bold
        ws[column[5] + self.row].alignment = align_top

        # ТАБЛИЦА
        def merge_table():
            ws.merge_cells(start_row=self.row, start_column=1, end_row=self.row, end_column=2)
            ws.merge_cells(start_row=self.row, start_column=3, end_row=self.row, end_column=5)
            ws[column[1] + self.row].font = font
            ws[column[3] + self.row].font = font
            ws[column[1] + self.row].alignment = align_top
            ws[column[3] + self.row].alignment = align_top
            b = 1
            while b < len(column):
                ws[column[b] + self.row].border = border
                b += 1

        def merge_table_bold():
            ws.merge_cells(start_row=self.row, start_column=1, end_row=self.row, end_column=2)
            ws.merge_cells(start_row=self.row, start_column=3, end_row=self.row, end_column=5)
            ws[column[1] + self.row].font = font_bold
            ws[column[3] + self.row].font = font_bold
            ws[column[1] + self.row].alignment = align_top
            ws[column[3] + self.row].alignment = align_top
            b = 1
            while b < len(column):
                ws[column[b] + self.row].border = border
                b += 1

        ws[column[1] + next_row()] = 'Дата'
        ws[column[3] + self.row] = 'Количество клиентов'
        merge_table_bold()
        # раскрвшивание фона для заголовков
        b = 1
        while b < len(column):
            ws[column[b] + self.row].fill = fill
            b += 1

        for line in client_count_totals_org:
            try:
                ws[column[1] + next_row()] = line[0].strftime('%d.%m.%Y')
                ws[column[3] + self.row] = line[1]
                merge_table()
            except AttributeError:
                pass

        ws[column[1] + next_row()] = 'Итого'
        ws[column[3] + self.row] = client_count_totals_org[-1][1]
        merge_table_bold()

        # увеличиваем все строки по высоте
        max_row = ws.max_row
        i = 2
        while i <= max_row:
            rd = ws.row_dimensions[i]
            rd.height = 18
            i += 1
        if client_count_totals_org[0][1]:
            date_ = datetime.strftime(client_count_totals_org[1][0], "%Y-%m")
        else:
            date_ = f'{datetime.strftime(client_count_totals_org[1][0], "%Y-%m-%d")} - ' \
                    f'{datetime.strftime(client_count_totals_org[-2][0], "%Y-%m-%d")}'
        path = self.local_folder + self.path + date_ + f' Количество клиентов за день по {client_count_totals_org[0][0]}' + ".xlsx"
        logging.info(f'{__name__}: {str(datetime.now())[:-7]}:    Сохранение отчета по количеству клиентов '
                     f'по {client_count_totals_org[0][0]} в {path}')
        path = self.create_path(path)
        self.save_file(path, wb)
        return path

    def import_xml_from_bitrix(self,
                               exchange_url,
                               exchange_path,
                               login,
                               password,
                               ):
        """
        Выгрузка новых заказов с сайта Битрикс
        :param exchange_url: адрес сайта
        :param exchange_path: путь к модулювыгрузки
        :param login: логин пользователя
        :param password: пароль
        :return: XML-файл с выгрузкой новых заказов
        """
        with requests.Session() as session:

            params = urllib.parse.urlencode({'type': 'sale', 'mode': 'checkauth'})
            url = 'http://' + exchange_url + exchange_path + '?' + params
            auth = (login, password)
            r = session.get(url, auth=auth)
            temp = r.text.split('\n')
            logging.info(str(temp))

            params = urllib.parse.urlencode({'type': 'sale', 'mode': 'init'})
            url = 'https://' + exchange_url + exchange_path + '?' + params
            result = session.get(url).text
            logging.info(result + '\n')

            params = urllib.parse.urlencode({'type': 'sale', 'mode': 'query'})
            url = 'http://' + exchange_url + exchange_path + '?' + params
            r = session.post(url)
            result = r.text
            logging.info(f'{datetime.strftime(datetime.now(), "%Y-%m-%d %H:%M:%S"):20}:    '
                  f'Запрос в битрикс выполнен.')
            if result[0] == '<':
                logging.info(
                    f'{datetime.strftime(datetime.now(), "%Y-%m-%d %H:%M:%S"):20}:    '
                    f'Данные выгружены, Отправка подтверждения в битрикс')
                params = urllib.parse.urlencode({'type': 'sale', 'mode': 'success'})
                url = 'http://' + exchange_url + exchange_path + '?' + params
                r = session.post(url)
            else:
                logging.error(f'{datetime.strftime(datetime.now(), "%Y-%m-%d %H:%M:%S"):20}:    Файл не содержит XML')

            return result

    def parseXML(self, xmlString):
        """
        ЧТЕНИЕ XML С ДАННЫМИ
        :param xmlString: строка XML из интернет-магазина Битрикс
        :return: список словарей с данными
        """

        x = re.search(r' encoding="windows-1251"', xmlString)
        xml = xmlString[:x.start()] + xmlString[x.end():]

        result = []
        products_in_bay = -1
        last_elem = ''

        root = objectify.fromstring(xml)

        for doc in root.getchildren():
            paydate = ''
            pay = False
            status = ''
            for req in doc.ЗначенияРеквизитов.getchildren():
                if req.Наименование == 'Дата оплаты':
                    paydate = datetime.strftime(datetime.strptime(str(req.Значение), '%d.%m.%Y %H:%M:%S'),
                                                '%Y-%d-%m %H:%M:%S')
                if req.Наименование == 'Заказ оплачен':
                    pay = bool(req.Значение)
                if req.Наименование == 'Статус заказа':
                    status = str(req.Значение)
            for product in doc.Товары.getchildren():
                count = int(product.Количество)
                while count > 0:
                    count -= 1
                    result.append(dict())
                    if last_elem != doc.Ид:
                        products_in_bay = -1
                    last_elem = doc.Ид
                    products_in_bay += 1
                    result[len(result) - 1]['Id_P'] = int(str(doc.Ид) + str(products_in_bay))
                    result[len(result) - 1]['OrderNumber'] = int(doc.Ид)
                    result[len(result) - 1]['ProductId'] = str(product.Ид)
                    result[len(result) - 1]['ProductName'] = str(product.Наименование)
                    result[len(result) - 1]['OrderDate'] = datetime.strptime(str(doc.Дата + ' ' + doc.Время),
                                                                             '%Y-%m-%d %H:%M:%S').strftime(
                        '%Y-%d-%m %H:%M:%S')
                    result[len(result) - 1]['PayDate'] = paydate
                    result[len(result) - 1]['Sum_P'] = Decimal(float(product.ЦенаЗаЕдиницу))
                    result[len(result) - 1]['Pay_P'] = pay
                    result[len(result) - 1]['Status_P'] = status
                    result[len(result) - 1]['Client_P'] = str(doc.Контрагенты.Контрагент.Ид)
        if result:
            logging.info(
                f'{datetime.strftime(datetime.now(), "%Y-%m-%d %H:%M:%S"):20}:    '
                f'Новый файл. Количество строк - {len(result)}')
        else:
            logging.info(f'{datetime.strftime(datetime.now(), "%Y-%m-%d %H:%M:%S"):20}:    Нет новых покупок.')
        return result

    def uploadToBase(self,
                     server,
                     database,
                     uid,
                     pwd,
                     Id_P,
                     OrderNumber_P,
                     ProductId_P,
                     ProductName_P,
                     OrderDate_P,
                     PayDate_P,
                     Sum_P,
                     Pay_P,
                     Status_P,
                     Client_P,
                     ):
        """
        Отправка данных в sql-базу
        """

        driver = '{SQL Server}'
        cnxn = pyodbc.connect(f'DRIVER={driver};SERVER={server};DATABASE={database};UID={uid};PWD={pwd}')
        cursor = cnxn.cursor()

        cursor.execute(f"""
                        INSERT INTO [Transactions](
                            [Id],
                            [OrderNumber],
                            [ProductId],
                            [ProductName],
                            [OrderDate],
                            [PayDate],
                            [Sum],
                            [Pay],
                            [Status],
                            [Client]
                        )
                        VALUES(
                            {Id_P},
                            {OrderNumber_P},
                            '{ProductId_P}',
                            '{ProductName_P}',
                            '{OrderDate_P}',
                            '{PayDate_P}',
                            {Sum_P},
                            {Pay_P},
                            '{Status_P}',
                            '{Client_P}'
                        )
                       """)
        cnxn.commit()
        return 'Upload To SQL-Base: Ready'

    def if_in_base(self,
                   server,
                   database,
                   uid,
                   pwd,
                   Id_P,
                   ):
        driver = '{SQL Server}'
        cnxn = pyodbc.connect(f'DRIVER={driver};SERVER={server};DATABASE={database};UID={uid};PWD={pwd}')
        cursor = cnxn.cursor()

        cursor.execute(f"""
                            SELECT
                                [Id],
                                [OrderNumber],
                                [ProductId],
                                [ProductName],
                                [OrderDate],
                                [PayDate],
                                [Sum],
                                [Pay],
                                [Status],
                                [Client]
                            FROM [Transactions]
                            WHERE
                                [Id] = {Id_P}
                           """)
        result = []
        while True:
            row = cursor.fetchone()
            if row:
                result.append(row)
            else:
                break
        if len(result) > 0:
            return False
        else:
            return True

    def request_bitrix(self):
        logging.info(
            f'\n--------------------------------------------------------------------------------------------------'
            f'----------------------------------------------------------\n'
            f'{datetime.strftime(datetime.now(), "%Y-%m-%d %H:%M:%S"):20}:    Запрос...')

        XML = self.import_xml_from_bitrix(exchange_url=self.bitrix_exchange_url,
                                          exchange_path=self.bitrix_exchange_path,
                                          login=self.bitrix_login,
                                          password=self.bitrix_password,
                                     )

        ordersList = self.parseXML(XML)

        for order in ordersList:

            if self.if_in_base(
                    server=self.server,
                    database=self.database_bitrix,
                    uid=self.user,
                    pwd=self.pwd,
                    Id_P=order['Id_P'],
            ):
                self.uploadToBase(
                    server=self.server,
                    database=self.database_bitrix,
                    uid=self.user,
                    pwd=self.pwd,
                    Id_P=order['Id_P'],
                    OrderNumber_P=order['OrderNumber'],
                    ProductId_P=order['ProductId'],
                    ProductName_P=order['ProductName'],
                    OrderDate_P=order['OrderDate'],
                    PayDate_P=order['PayDate'],
                    Sum_P=order['Sum_P'],
                    Pay_P=int(order['Pay_P']),
                    Status_P=order['Status_P'],
                    Client_P=order['Client_P'],
                )

                logging.info(
                    f'{datetime.strftime(datetime.now(), "%Y-%m-%d %H:%M:%S"):20}:    '
                    f'{order["ProductName"]:50} | {order["OrderNumber"]:10} | '
                    f'{order["Sum_P"]: 10} | {order["OrderDate"]:25} | ВЫГРУЖЕН')
            else:
                logging.info(
                    f'{datetime.strftime(datetime.now(), "%Y-%m-%d %H:%M:%S"):20}:    '
                    f'{order["ProductName"]:50} | {order["OrderNumber"]:10} | '
                    f'{order["Sum_P"]: 10} | {order["OrderDate"]:25} | ОТМЕНА (Попытка повторной записи)')

    def load_checkbox(self):
        """
        Установка чекбоксов в соответствии с настройками INI-файла
        """
        logging.info(f'{__name__}: {str(datetime.now())[:-7]}:    Загрузка настроек...')
        self.root.ids.report.ids.split_by_days.active = self.split_by_days
        self.root.ids.report.ids.finreport_xls.active = self.finreport_xls
        self.root.ids.report.ids.check_client_count_total_xls.active = self.check_client_count_total_xls
        self.root.ids.report.ids.check_cashreport_xls.active = self.check_cashreport_xls
        self.root.ids.report.ids.check_itogreport_xls.active = self.check_itogreport_xls
        self.root.ids.report.ids.agentreport_xls.active = self.agentreport_xls
        self.root.ids.report.ids.use_yadisk.active = self.use_yadisk
        self.root.ids.report.ids.finreport_google.active = self.finreport_google
        self.root.ids.report.ids.finreport_telegram.active = self.finreport_telegram

    def change_checkbox(self, name, checkbox):
        """
        Изменяет состояние элемента конфигурации и записывает в INI-файл
        :param name: Имя чекбокса
        :param checkbox: Состояние active чекбокса
        """
        self.config.set('General', name, str(checkbox))
        setattr(self, name, checkbox)
        self.config.write()
        logging.info(f'{__name__}: {str(datetime.now())[:-7]}:    Параметр {name} изменен на значение {checkbox}')
        if name == 'split_by_days' and not checkbox and not self.root.ids.report.ids.date_switch.active:
            self.root.ids.report.ids.finreport_google.active = False
            self.change_checkbox('finreport_google', False)
            self.root.ids.report.ids.finreport_google.disabled = True
            self.root.ids.report.ids.finreport_google_text.disabled = True
        elif name == 'split_by_days' and checkbox:
            self.root.ids.report.ids.finreport_google_text.disabled = False
            self.root.ids.report.ids.finreport_google.disabled = False

    def save_reports(self):
        """
        Функция управления
        """
        self.fin_report()
        self.agent_report()
        if self.finreport_xls:
            self.path_list.append(self.export_fin_report())
        if self.agentreport_xls:
            self.path_list.append(self.export_agent_report(self.agentreport_dict))
        if self.finreport_google:
            self.request_bitrix()
            self.fin_report_lastyear()
            self.export_to_google_sheet()
            self.open_googlesheet()
        if self.finreport_telegram:
            self.sms_report_list.append(self.sms_report())
        if self.check_itogreport_xls:
            if self.itog_report_org1['Итого по отчету'][1]:
                self.path_list.append(self.save_organisation_total(self.itog_report_org1))
            if self.itog_report_org2['Итого по отчету'][1]:
                self.path_list.append(self.save_organisation_total(self.itog_report_org2))
        if self.check_cashreport_xls:
            if self.cashdesk_report_org1['Итого'][0][1]:
                self.path_list.append(self.save_cashdesk_report(self.cashdesk_report_org1))
            if self.cashdesk_report_org2['Итого'][0][1]:
                self.path_list.append(self.save_cashdesk_report(self.cashdesk_report_org2))
        if self.check_client_count_total_xls:
            if self.client_count_totals_org1[-1][1]:
                self.path_list.append(self.save_client_count_totals(self.client_count_totals_org1))
            if self.client_count_totals_org2[-1][1]:
                self.path_list.append(self.save_client_count_totals(self.client_count_totals_org2))

    def load_report(self):
        """
        Выполнить отчеты
        """
        self.itog_report_org1 = None
        self.itog_report_org2 = None
        self.report_bitrix = None

        self.click_select_org()

        if self.org1:
            self.itog_report_org1 = self.itog_report(
                server=self.server,
                database=self.database1,
                driver=self.driver,
                user=self.user,
                pwd=self.pwd,
                org=self.org1[0],
                org_name=self.org1[1],
                date_from=self.date_from,
                date_to=self.date_to,
                hide_zeroes='0',
                hide_internal='1',
            )
            self.itog_report_org1_lastyear = self.itog_report(
                server=self.server,
                database=self.database1,
                driver=self.driver,
                user=self.user,
                pwd=self.pwd,
                org=self.org1[0],
                org_name=self.org1[1],
                date_from=self.date_from - relativedelta(years=1),
                date_to=self.date_to - relativedelta(years=1),
                hide_zeroes='0',
                hide_internal='1',
            )
            self.cashdesk_report_org1 = self.cashdesk_report(
                server=self.server,
                database=self.database1,
                driver=self.driver,
                user=self.user,
                pwd=self.pwd,
                date_from=self.date_from,
                date_to=self.date_to,
            )
            self.client_count_totals_org1 = self.client_count_totals_period(
                server=self.server,
                database=self.database1,
                driver=self.driver,
                user=self.user,
                pwd=self.pwd,
                org=self.org1[0],
                org_name=self.org1[1],
                date_from=self.date_from,
                date_to=self.date_to,
            )
        if self.org2:
            self.itog_report_org2 = self.itog_report(
                server=self.server,
                database=self.database2,
                driver=self.driver,
                user=self.user,
                pwd=self.pwd,
                org=self.org2[0],
                org_name=self.org2[1],
                date_from=self.date_from,
                date_to=self.date_to,
                hide_zeroes='0',
                hide_internal='1',
            )
            self.cashdesk_report_org2 = self.cashdesk_report(
                server=self.server,
                database=self.database2,
                driver=self.driver,
                user=self.user,
                pwd=self.pwd,
                date_from=self.date_from,
                date_to=self.date_to,
            )
            self.client_count_totals_org2 = self.client_count_totals_period(
                server=self.server,
                database=self.database2,
                driver=self.driver,
                user=self.user,
                pwd=self.pwd,
                org=self.org2[0],
                org_name=self.org2[1],
                date_from=self.date_from,
                date_to=self.date_to,
            )
        self.report_bitrix = self.read_bitrix_base(
            server=self.server,
            database=self.database_bitrix,
            user=self.user,
            pwd=self.pwd,
            driver=self.driver,
            date_from=self.date_from,
            date_to=self.date_to,
        )
        self.report_bitrix_lastyear = self.read_bitrix_base(
            server=self.server,
            database=self.database_bitrix,
            user=self.user,
            pwd=self.pwd,
            driver=self.driver,
            date_from=self.date_from - relativedelta(years=1),
            date_to=self.date_to - relativedelta(years=1),
        )
        # Чтение XML с привязкой групп услуг к услугам
        self.orgs_dict = self.read_reportgroup(self.reportXML)
        # Поиск новых услуг
        self.find_new_service(self.itog_report_org1, self.orgs_dict)
        self.distibution_service()

    def run_report(self):
        self.open_browser = False
        self.path_list = []
        self.sms_report_list = []

        if self.date_switch:
            self.load_report()
            pass
        else:
            if self.split_by_days:
                period = []
                while True:
                    period.append(self.date_from)
                    if self.date_from + timedelta(1) == self.date_to:
                        break
                    else:
                        self.date_from = self.date_from + timedelta(1)
                for date in period:
                    self.date_from = date
                    self.date_to = date + timedelta(1)
                    self.load_report()
                self.date_from = datetime.strptime(self.root.ids.report.ids.date_from.text, "%Y-%m-%d")
            else:
                self.load_report()
        # Отправка в яндекс диск
        if self.use_yadisk:
            self.sync_to_yadisk(self.path_list, self.yadisk_token)
            self.path_list = []
        # Отправка в телеграмм
        if self.finreport_telegram:
            self.send_message_to_telegram()
            self.sms_report_list = []

if __name__ == '__main__':
    pass
